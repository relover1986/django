#%%
import pyecharts.options as opts
from pyecharts.charts import Bar, Grid, Line
from pyecharts.charts import Line,Grid,Page,Timeline,Bar
from pyecharts.charts import Page, Sankey,Graph
import numpy as np
import pandas as pd 
from pyecharts.components import Table
from pyecharts.options import ComponentTitleOpts
from numpy import mean
from PIL import Image  #导入PIL库
from django.shortcuts import redirect, render,HttpResponse
from io import BytesIO
from itertools import permutations
import random
import datetime
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
from aip import AipOcr
from django.conf import settings
import os
# 定义常量
APP_ID = '23933903'
API_KEY = 'n8UVViIuLYnmRfKAzq5z5rvf'
SECRET_KEY = 'vEowelgK61XIFbiinXITEZLMnno9HSsv'

# 初始化AipFace对象
aipOcr = AipOcr(APP_ID, API_KEY, SECRET_KEY)
options = {}
options["detect_direction"] = "true"  # 检测朝向
options["detect_language"] = "true"  # 检测语言

base_dir =settings.MEDIA_ROOT    





def 合同(old_text, new_text,document):
    all_paragraphs = document.paragraphs
    for paragraph in all_paragraphs:
        for run in paragraph.runs:
            run_text = run.text.replace(old_text, new_text)
            run.text = run_text



    all_tables = document.tables
    for table in all_tables:
            for row in table.rows:
                for cell in row.cells:
                    cell_text = cell.text.replace(old_text, new_text)
                    cell.text = cell_text



























 
 
def find_chinese(file):
    pattern = re.compile(r'[^\u4e00-\u9fa5]')
    chinese = re.sub(pattern, '', file)
    return chinese
 
def find_unchinese(file):
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    unchinese = re.sub(pattern,"",file)
    print(unchinese)
 
 
 


def df_max(data,col):
    df=pd.DataFrame(data)
    df[col]=df[col].astype('int')
    return df[col].max()

def DataFrame(data):
    df=pd.DataFrame(data)
    return df
def dframe(data):
    df=pd.DataFrame(data)
    lst=list(df.columns)

    return lst

def run_dframe(data):
    df=pd.DataFrame(data)
    df['配速']=(df['分钟']+(df['秒']/60))/df['距离']
    lst=list(df.columns)

    return lst,df.values()
    


def df_ident(data,ident):
    df=pd.DataFrame(data)
    ident=df[ident].astype(int).max()+1

    return (6-len(str(ident)))*'0'+str(ident)



def add_list(fun):
    print(fun.__name__)
    def check(request):


        if "000001" in request.session.get('info').get('ident'):
            return fun(request)
        elif  request.session.get('info').get('role') in fun.__name__:
            return fun(request)
        else:
            error_msg = "没有权限!"
            return HttpResponse(error_msg)        
    return check


def url_check(fun):
    print(fun.__name__)
    def check(request):


        if "000001" in request.session.get('info').get('ident'):
            return fun(request)
        elif  "教练" in request.session.get('info').get('name'):
            return fun(request)
        else:
            error_msg = "没有权限!"
            return redirect('/home')        
    return check

def 最高权限(fun):
    print(fun.__name__)
    def check(request):
        # 修改条件判断逻辑，支持多身份验证
        if any(x in request.session.get('info').get('ident') for x in ["000001", "000002"]):
            return fun(request)
        else:
            title="没有权限!"
            return render(request,'change.html',{"title":title })  
    return check

from app01.permissions import 资料员


def tu_128(file,username):

    hou=os.path.splitext(str(file))[-1]
    print(hou)

    if hou in ['.png','.jpg','.jpeg']:

        path = './app01/static/img/'
        image_name = username+'.jpg'
        f_path=path+image_name
        with open(f_path, mode='wb') as f:
            for i in file.chunks():
                f.write(i)

        img = Image.open(f_path)  # 读取图片
        w, h = img.size  # 输出图片(宽度w,高度h)

        倍数 = w / 128

        if w > 128:
            
            img = img.resize((128, int(h / 倍数)), Image.ANTIALIAS)

            img.save(f_path)
# 

def line_range(df,图,lable,x_data,y_data):
    
    lst=list(set(df[图].tolist()))
    lst.sort()

    c=locals()
    page=Page(layout=Page.SimplePageLayout)

    for  i in lst:
        df_=df[df[图]==i]


        c[str(i)] = (
        Line()     
        .set_global_opts(title_opts=opts.TitleOpts(title=i, pos_top=1,),
                        legend_opts=opts.LegendOpts(pos_top=30)
        ))
        c[str(i)].add_xaxis(xaxis_data=x_data) 

        for l in df_[lable].tolist():

            

            df1=df_[df_[lable]==l]



            

                     
            
            c[str(i)].add_yaxis(
                series_name=str(l),
                # symbol=str(i),
                y_axis=df1[y_data].tolist(),
                is_smooth=True,
                # label_opts=opts.LabelOpts(formatter=str(i))
                
                )
            
        (Grid(init_opts=opts.InitOpts(width="1000px", height="1920px"))
                .add(chart=c[str(i)], grid_opts=opts.GridOpts(pos_left=0, pos_top=150, height="35%")))
                
        page.add(c[str(i)])


    return page


#line_n(df,'动作','项目','日期','平均心率')
def line_n(df,图,lable,x_data,y_data):

    lst=list(set(df[图].tolist()))
    lst.sort()

    c=locals()
    page=Page(layout=Page.SimplePageLayout)
    for  i in lst:
        df_=df[df[图]==i]


        c[str(i)] = (
        Line()     
        .set_global_opts(title_opts=opts.TitleOpts(title=i, pos_top=1,),
                        legend_opts=opts.LegendOpts(pos_top=30)
        ))
        c[str(i)].add_xaxis(xaxis_data=df_[x_data].tolist()) 

        for l in df_[lable].tolist():

            

            df1=df_[df_[lable]==l]
            df_y=df_[['日期']].merge(df1,on='日期',how='left')
            

                     
            
            c[str(i)].add_yaxis(
                series_name=str(l),
                # symbol=str(i),
                y_axis=df_y[y_data].tolist(),
                is_smooth=True,
                # label_opts=opts.LabelOpts(formatter=str(i))
                
                )
            
        (Grid(init_opts=opts.InitOpts(width="1280px", height="720px"))
                .add(chart=c[str(i)], grid_opts=opts.GridOpts(pos_left=0, pos_top=150, height="35%")))
                
        page.add(c[str(i)])


    return page



def hr1(data):
    df=pd.DataFrame(data)
    df=df.copy()
    df=df.drop_duplicates(subset='日期',keep="last")

    data=[]
    for i in df.index:
        df_=df[df.index==i].copy()
        df_['心率']=df_['心率'].str.split('.')
        df_1=df_.explode('心率',ignore_index=True)

        data.append(df_1)
    df=pd.concat(data)
    df=df.reset_index()
    df['index']=df['index'].astype('int')


    #


    df['时间']=(df['index'])*df['间隔时间'].astype('int')



    df['项目']=df['动作']+'-'+df['重量'].astype(str)+'-'+df['间隔时间'].astype(str)
    df=df[['日期','动作','项目','时间','心率']]
    df['日期']=df['日期'].astype('datetime64[ns]')
    df['日期']=df['日期'].dt.strftime('%Y-%m-%d')
    df=df.sort_values(by=['日期','时间'])
    df['时间']=df['时间'].astype(str)
    return df
def hr(data):
    df=pd.DataFrame(data)

    df=df.copy()
    df=df.drop_duplicates(subset='日期',keep="last")

    df.loc[:,'平均心率']=df['心率'].str.split('.').apply(lambda x: mean([int(i) for i in x])).astype('int')


    df=df.reset_index()

    df['项目']=df['动作']+'-'+df['重量'].astype(str)+'-'+df['间隔时间'].astype(str)
    df=df[['日期','项目','平均心率']]


    df['日期']=df['日期'].astype('datetime64[ns]')
    df=df.sort_values(by=['日期'])
    df['日期']=df['日期'].dt.strftime('%Y-%m-%d')
    df['动作']=df['项目'].str.split(pat='-').str[0]
    df=df.sort_values(by=['日期','项目'])
    return df

def meituan(df,ziduan):
    c =   Bar() .add_xaxis(df[ziduan].tolist()
        ) .add_yaxis("问题单", df['数量'].tolist()).set_global_opts(
            xaxis_opts=opts.AxisOpts(axislabel_opts=opts.LabelOpts(rotate=-15)),
            title_opts=opts.TitleOpts(title="Bar", subtitle=""),
        )
    return c

def meituan_df(data,ziduan):
    df=pd.DataFrame(data)

    df=df.copy()
    df['数量']=1
    df=df[['城市','团队名称','问题单类型','数量']]
    df1=df.groupby([ziduan]).sum().reset_index()
    return df1

def meituan2(data,txt):
    df=pd.DataFrame(data)

    df=df.copy()
    df['数量']=1

    
    df=df[['团队名称','骑手姓名',	'骑手手机号', '问题单类型','数量']]
    df=df.groupby(['团队名称','骑手姓名',	'骑手手机号', '问题单类型']).sum().reset_index()

    df=df.sort_values(by='数量',ascending=False)





    table = Table()

    headers = list(df.columns)
    rows = df.values
    table.add(headers, rows)
    table.set_global_opts(
        title_opts=ComponentTitleOpts(title=txt, subtitle="")
    )
    return table


def lst():
    a = random.randint(1,13)
    b = random.randint(1,13)     
    c = random.randint(1,13)
    d = random.randint(1,13)
    my_list = [a, b, c, d]
    # 对4个整数随机排列的列表
    result = [c for c in permutations(my_list, 4)]

    symbols = ["+", "-", "*", "/"]

    list2 = []  # 算出24的排列组合的列表

    flag = False


    for one, two, three, four in result:
        for s1 in symbols:
            for s2 in symbols:
                for s3 in symbols:
                    if s1 + s2 + s3 == "+++" or s1 + s2 + s3 == "***":
                        express = ["{0}{1}{2}{3}{4}{5}{6}".format(one, s1, two, s2, three, s3, four)]  # 全加或者乘时，括号已经没有意义。
                    else:
                        express = ["(({0}{1}{2}){3}{4}){5}{6}".format(one, s1, two, s2, three, s3, four),
                                "({0}{1}{2}){3}({4}{5}{6})".format(one, s1, two, s2, three, s3, four),
                                "(({0}{1}({2}{3}{4})){5}{6})".format(one, s1, two, s2, three, s3, four),
                                "{0}{1}(({2}{3}{4}){5}{6})".format(one, s1, two, s2, three, s3, four),
                                "{0}{1}({2}{3}({4}{5}{6}))".format(one, s1, two, s2, three, s3, four)]
                    # print(one + two + three + four)
                        if str(one) + str(two) + str(three) + str(four) == "8383":
                            #print(express)
                            pass

                    for e in express:
                        try:
                            # if eval(e) == 24:
                            if round(eval(e), 6) == 24:
                                list2.append(e)
                                flag = True
                        except ZeroDivisionError:
                            pass

    list3 = set(list2) # 去除重复项

    if list3:
    # print(my_list)
        return my_list
    else:
        return 
        

def Textcolor(file):
    wb = load_workbook(file) 

    #Color=['c6efce','006100']#绿
    #Color = ['ffc7ce', '9c0006']  #红
    Color = ['ffeb9c', '9c6500']  # 黄
    #Color = ['ffffff', '000000']  # 黑白

    fille = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
    font = Font(u'微软雅黑', size=18, bold=True, italic=False, strike=False, color=Color[1])  # 设置字体样式

    ws = wb.worksheets[0]
    col_N=ws.max_column
    row_N=ws.max_row 
    ws.cell(row_N+1,col_N).value=datetime.date.today()
    ws.cell(row_N+1,col_N-1).value='日期'

    for c in range(col_N):
        列=get_column_letter(c+1)
        ws.column_dimensions[列].width = 21#修改列D的列宽
        for i in range(row_N-1):
            ws.cell(row=i + 2, column=c+1 ).font =font # 序列
            if i%2==1:

            

                ws.cell(row=i + 2, column=c+1).fill = fille  # 序列


    wb.save(file) 


def No_24(m):

    表=[]
    for i in range(2):

        data=[]
        n=0
        while True:
            
            a=lst()
            
            if a:
                df=pd.DataFrame({'数字':','.join([str(i) for i in a]),'书写':''},index=[0])
                data.append(df)
                n+=1
            if n==m:
                break

        df=pd.concat(data)
        表.append(df)

    df=pd.concat(表,axis=1)
    file = BytesIO()
    df.to_excel(file,index=False)
    Textcolor(file)
    file.seek(0)
    return df,file


def table(df,txt):
    table = Table()

    headers = list(df.columns)
    if len(headers)!=len(set(headers)):
        headers=list(range(len(headers)))

    rows = df.values
    table.add(headers, rows)
    table.set_global_opts(
        title_opts=ComponentTitleOpts(title=txt, subtitle="")
    )
    return table



def Answer(a,b,c,d):

    my_list = [a, b, c, d]
    # 对4个整数随机排列的列表
    result = [c for c in permutations(my_list, 4)]

    symbols = ["+", "-", "*", "/"]

    list2 = []  # 算出24的排列组合的列表

    flag = False


    for one, two, three, four in result:
        for s1 in symbols:
            for s2 in symbols:
                for s3 in symbols:
                    if s1 + s2 + s3 == "+++" or s1 + s2 + s3 == "***":
                        express = ["{0}{1}{2}{3}{4}{5}{6}".format(one, s1, two, s2, three, s3, four)]  # 全加或者乘时，括号已经没有意义。
                    else:
                        express = ["(({0}{1}{2}){3}{4}){5}{6}".format(one, s1, two, s2, three, s3, four),
                                   "({0}{1}{2}){3}({4}{5}{6})".format(one, s1, two, s2, three, s3, four),
                                   "(({0}{1}({2}{3}{4})){5}{6})".format(one, s1, two, s2, three, s3, four),
                                   "{0}{1}(({2}{3}{4}){5}{6})".format(one, s1, two, s2, three, s3, four),
                                   "{0}{1}({2}{3}({4}{5}{6}))".format(one, s1, two, s2, three, s3, four)]
                    # print(one + two + three + four)
                        if str(one) + str(two) + str(three) + str(four) == "8383":
                            #print(express)
                            pass

                    for e in express:
                        try:
                            # if eval(e) == 24:
                            if round(eval(e), 6) == 24:
                                list2.append(e)
                                flag = True
                        except ZeroDivisionError:
                            pass

    list3 = set(list2) # 去除重复项
    data=[]
    n=1
    for i in list3:

        df=pd.DataFrame({'序号':str(n),'答案':i},index=[0])
        data.append(df)
        n+=1
    df=pd.concat(data).head(3)
    return df



def txt2mp3(file,filename):

    hou=os.path.splitext(str(file))[-1]
    print(hou)
    file_path = os.path.join(base_dir, str(file))

    if hou in ['.txt']:
        with open (file_path,'r',encoding='utf-8') as f:
            txt=f.read()
            txt=find_chinese(txt)
            txt=txt.replace(" ", "。")
            txt=txt.replace("\n", "。。。。")
    elif hou in ['.jpg','.png','.jpeg']:
        txt=Ocr(file_path)
        txt=txt.replace("\n", "。。。。")
        print(txt)

        
    file_name=os.path.join(base_dir, 'mp3/'+str(filename)+'.mp3')

    os.system('edge-tts --voice zh-CN-liaoning-XiaobeiNeural --text {} --write-media {}'.format(txt,file_name))   
    return txt




# 打开图片
def get_file_content(filePath):
    with open(filePath, 'rb') as fp:
        return fp.read()


# 通用文字识别（高精度版）
def 高精度(file):
    options = {}
    options["detect_direction"] = "true"  # 检测朝向
    options["detect_language"] = "true"  # 检测语言
    result = aipOcr.basicAccurate(file, options)
    return (result)



def Ocr(图片):


    file = get_file_content(图片)
    result = 高精度(file)

    txt = ''
    for word in result['words_result']:

        txt = txt + find_chinese(word['words'])

    return txt

def Safe_sankey(df):
    表=[]
    col=len(df.columns)
    for i in range(col-2):
    

        data=df[[df.columns[i],df.columns[i+1],df.columns[col-1]]]
        data=data.groupby([df.columns[i],df.columns[i+1]],as_index=False).sum()
        data.columns=['0','1','2']    
        表.append(data)

    data=pd.concat(表)
    data.columns=[1,0,2]

    # 生成nodes
    nodes = []
    nodes.append({'name':data[1].tolist()[0]})

    for i in data[0].unique():
        dic = {}
        dic['name'] = i
        nodes.append(dic)
    
    # 生成links
    links = []
    for i in data.values:
        dic = {}
        dic['source'] = i[0]
        dic['target'] = i[1]
        dic['value'] = i[2]
        links.append(dic)


    c = (
            Sankey(init_opts=opts.InitOpts(width="1000px", height="900px"))
            .add(
    "风险因素",
                nodes,
                links,
                linestyle_opt=opts.LineStyleOpts(opacity=0.2, curve=0.5, color="source",type_="dotted"),
                label_opts=opts.LabelOpts(position="right",),
            )
            .set_global_opts(title_opts=opts.TitleOpts(title="安全风险因素"))
        )
    # 输出html可视化结果
    return c


def guanxitu(data,源):
    
    df=pd.DataFrame(data) 
    df_links=df.copy()   
    if 源:
        # df=df[df['name']==源]
        df_links=df[df['source']==源]

        df1=df[df['source'].isin(df_links['target'].tolist())]
        df_links=pd.concat([df_links,df1])
    

    links=df.to_dict('records')
    data=[]
    for i in df['source'].unique():
        df2=df[df['source']==i]
        df2.columns=['category','name']

        df2['symbolSize']=5
        df2['draggable']=False
        df2['value']=0#len(df2)

        dic={}
        dic['category']=i
        dic['name']=i
        dic["label"] ={'normal': {'show': True}}
        dic['symbolSize']=len(df2)/10
        dic['draggable']=False
        dic['value']=len(df2)
        df_=pd.DataFrame(dic,index=[0])

        data.append(df2)
        data.append(df_)
    df=pd.concat(data)
    df1=df[(df['name']==df['category'])]
    df2=df[~(df['name']==df['category'])]
    df=pd.concat([df1,df2])
    df=df.drop_duplicates(subset=['name'])


    df_nodes=df.copy()
    df_=df_links.copy()


  



    data_array = np.array(df_.values)
    data_list =data_array.tolist()
    list_2 = sum(data_list, [])
    df_name=pd.DataFrame({'name':list_2 })#DataFrame
    df_name=df_name.drop_duplicates(subset=['name'],keep='first')

    # df_=df_name.copy()
    # df_.columns=['source']

    categories=df_name.to_dict('records')
    df_nodes=pd.merge(df_name,df_nodes,how='left',on='name')

    nodes=df_nodes.to_dict('records')



    c = (
        Graph(init_opts=opts.InitOpts(width="1200px", height="900px"))
            
        .add(
            "",
            nodes,
            links,
            categories,
            repulsion=50,
            linestyle_opts=opts.LineStyleOpts(curve=0.2),
            label_opts=opts.LabelOpts(is_show=False),
        )
        .set_global_opts(
            legend_opts=opts.LegendOpts(is_show=False),
            title_opts=opts.TitleOpts(title=源+"关系图"),
        )
        
    )
    return c.render_embed()

def run_line(df):
    x_data = df['日期'].tolist()
    bar = (
        Bar()
        .add_xaxis(x_data)
        .add_yaxis(
            "距离",
            df['距离'].tolist(),
            yaxis_index=0,
            color="#d14a61",
        )
        .add_yaxis(
            "平均心率",
            df['平均心率'].tolist(),
            yaxis_index=1,
            color="#5793f3",
        )
        .extend_axis(
            yaxis=opts.AxisOpts(
                name="心率",
                type_="value",
                min_=100,
                max_=220,
                position="right",
                axisline_opts=opts.AxisLineOpts(
                    linestyle_opts=opts.LineStyleOpts(color="#d14a61")
                ),
                axislabel_opts=opts.LabelOpts(formatter="{value} 次/分钟"),
            )
        )
        .extend_axis(
            yaxis=opts.AxisOpts(
                type_="value",
                name="配速",
                min_=3,
                max_=7,
                position="left",
                axisline_opts=opts.AxisLineOpts(
                    linestyle_opts=opts.LineStyleOpts(color="#675bba")
                ),
                axislabel_opts=opts.LabelOpts(formatter="{value}分"),
                splitline_opts=opts.SplitLineOpts(
                    is_show=True, linestyle_opts=opts.LineStyleOpts(opacity=1)
                ),
            )
        )
        .set_global_opts(
            yaxis_opts=opts.AxisOpts(
                name="距离",
                min_=0,
                max_=10,
                position="right",
                offset=80,
                axisline_opts=opts.AxisLineOpts(
                    linestyle_opts=opts.LineStyleOpts(color="#5793f3")
                ),
                axislabel_opts=opts.LabelOpts(formatter="{value}公里"),
            ),
            title_opts=opts.TitleOpts(title=""),
            tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),
        )
    )

    line = (
        Line()
        .add_xaxis(x_data)
        .add_yaxis(
            "配速",
            df['配速'].tolist(),
            yaxis_index=2,
            color="#675bba",
            label_opts=opts.LabelOpts(is_show=False),
        )
    )

    bar.overlap(line)
    grid = Grid()
    grid.add(bar, opts.GridOpts(pos_left="5%", pos_right="20%"), is_control_axis_index=True)

    return grid.render_embed()


def fen (file,heng,shu):   

    path = './media/'
    file=path+str(file)
    hou=os.path.splitext(str(file))[-1].lower() 
    print(file,heng,shu) 
    if hou in ['.png','.jpg','.jpeg']:

        img = Image.open(file)  # 读取图片
       
        width, height = img.size

        w=random.randint(1,heng)
        h=random.randint(1,shu)
        print(w,h) 
        for i in range(shu):
            for j in range(heng):
                if h==i and w==j:
                    print(i,j)
                    box = (j * width // heng, i * height // shu, (j + 1) * width // heng, (i + 1) * height // shu)
                    crop_img = img.crop(box)
                    crop_img.save(file)

def bpzd (k,a,q,v,r):  
    k=float(k)
    a=float(a)


    if len(v) == 0:
        q=float(q)
        r=float(r)    
        v=k*(q**(1/3)/r)**a
        v=round(v, 2)
    elif len(q) == 0:
        v=float(v)
        r=float(r)
        q=((v/k)**(1/a)*r)**3
        q=round(q, 2)
    elif len(r) == 0:
        v=float(v)
        q=float(q)
        r=q**(1/3)/((v/k)**(1/a))
        r=round(r, 2)


    df=pd.DataFrame({'k':str(k),'α':a,'v':v,'q':q,'r':r},index=[0])

    return df