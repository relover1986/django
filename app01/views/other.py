from django.shortcuts import render, HttpResponse, redirect
import time
import torch
import torch.nn as nn
import json
import tempfile
import cv2
from rapidocr_onnxruntime import RapidOCR
from django.http import JsonResponse
from app01 import models
from app01.models import UploadedZhaopian
from app01 import modelform
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_protect, csrf_exempt
import zipfile
import io
from collections import defaultdict
from app01.jiami import md5
from app01.func import *
from app01.photo import *
from django_filters.views import FilterView
from django.views.generic import ListView
import pandas as pd
from openpyxl import load_workbook
import os
from docx import Document
import sqlite3 as sl
from django.shortcuts import render
from django.apps import apps
from app01.models import ExplosiveInventoryItem
from app01.modelform import ExplosiveInventoryItemForm
from PIL import Image
import io
import os
from django.core.files.base import ContentFile
from openpyxl import Workbook
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage
import base64
from django.core.validators import MinLengthValidator, MaxLengthValidator, RegexValidator
from django.conf import settings
import uuid
import json
import sys
sys.path.insert(0, "/root/django")
import utils
import sys
sys.path.insert(0, "/root/django")
import utils

from openpyxl import Workbook
from django.http import HttpResponse
import io
from datetime import datetime


import zipfile
from io import BytesIO

from aip import AipBodyAnalysis
import base64
import io
from PIL import Image
import os
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework import status
# 百度API配置
APP_ID = '118049497'
API_KEY = 'AmK3oZpZhns9jAm2rJgzRyLq'
SECRET_KEY = 'LbbCCzQyv1FlytQxBHstZ5Yt5i4B7pMw'
client = AipBodyAnalysis(APP_ID, API_KEY, SECRET_KEY)


def tu_add(request):
    title = 'tu'
    if request.method == 'POST':

        model_name = request.POST['model_name']
        files = request.FILES.getlist('file')  # 获取所有上传的文件

        for file in files:
            time.sleep(0.1)
            # 处理每个文件，例如保存到数据库或文件系统
            models.UploadedTu.objects.create(
                model_name=model_name, pdf_file=file)

        return redirect("/home/tu_list")

    else:
        model_names = models.UploadedTu.objects.values_list(
            'model_name', flat=True).distinct()
        return render(request, 'tu_add.html', {'model_names': model_names})


def tu_delete(request):

    id = request.GET.get('id')
    models.UploadedTu.objects.filter(id=str(id)).delete()
    return redirect("/home/tu_list")


def tu_list(request):

    title = 'tu'
    if request.method == "GET":

        data = models.UploadedTu.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at")[:100]

        # print(data)

        lst = dframe(data)
        cols = []

        for i in lst:
            cols.append({'age': i})

        return render(request, 'tu_list.html', {"data": data, "cols": cols, "title": title})

# %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%


def pdf_add(request):
    title = 'tu'
    if request.method == 'POST':

        model_name = request.POST['model_name']
        files = request.FILES.getlist('file')  # 获取所有上传的文件

        for file in files:
            time.sleep(0.1)
            # 处理每个文件，例如保存到数据库或文件系统
            models.UploadedPDF.objects.create(
                model_name=model_name, pdf_file=file)

        return redirect("/home/pdf_list")

    else:
        model_names = models.UploadedPDF.objects.values_list(
            'model_name', flat=True).distinct()
        return render(request, 'pdf_add.html', {'model_names': model_names})


def pdf_delete(request):

    id = request.GET.get('id')
    models.UploadedPDF.objects.filter(id=str(id)).delete()
    return redirect("/home/pdf_list")


def pdf_list(request):

    title = 'pdf'
    if request.method == "GET":

        data = models.UploadedPDF.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at")[:100]

        # print(data)

        lst = dframe(data)
        cols = []

        for i in lst:
            cols.append({'age': i})

        return render(request, 'pdf_list.html', {"data": data, "cols": cols, "title": title})


# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################


def inventory_list(request):

    title = 'inventory'
    database = '出入库记录'

    if request.method == "GET":

        model_fields = models.ExplosiveInventoryItem._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields if field.attname not in ('id', 'location')]
        cols.append({'verbose_name': '操作'})
        data = models.ExplosiveInventoryItem.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at").order_by(
            '-date')[:100]

        return render(request, 'inventory_list.html', {"data": data, "cols": cols, "数据库": database, 'title': title, "export_xlsx_url": "/home/inventory_export_xlsx"  # 新增XLSX导出参数
                                                       })


def inventory_add(request):
    title = '出入库记录'

    if request.method == "GET":

        form = ExplosiveInventoryItemForm()

        return render(request, 'card_form.html', {'form': form, '标题': title})

    if request.method == 'POST':
        form = ExplosiveInventoryItemForm(request.POST)

        if form.is_valid():

            form.save()

        else:

            form.errors
            return render(request, 'card_form.html', {'form': form, '标题': title})

    return redirect("/home/inventory_list")


def inventory_delete(request):

    id = request.GET.get('id')
    models.ExplosiveInventoryItem.objects.filter(id=str(id)).delete()
    return redirect("/home/inventory_list")


def inventory_edit(request):

    title = '出入库记录'
    id = request.GET.get('id')
    row_object = models.ExplosiveInventoryItem.objects.filter(
        id=str(id)).first()

    if request.method == "GET":

        form = modelform.ExplosiveInventoryItemForm(instance=row_object)

        return render(request, 'card_form.html', {"form": form, "标题": title})

    form = modelform.ExplosiveInventoryItemForm(
        data=request.POST, instance=row_object)

    if form.is_valid():

        form.save()
    else:
        title = '输入错误'
        form.errors
        return render(request, 'card_form.html', {'form': form})
    return redirect("/home/inventory_list")


def inventory_export_xlsx(request):

    # 获取数据
    inventory_items = models.ExplosiveInventoryItem.objects.all().values('id',
                                                                         'project_department', 'blaster', 'emulsion_explosive_32mm',
                                                                         'powdery_explosive_box_2', 'sticky_explosive', 'electronic_detonator_5m',
                                                                         'electronic_detonator_15m', 'inventory_status', 'detonating_device_quantity',
                                                                         'detonating_cord_length', 'date'
                                                                         )

    # 将QuerySet转换为pandas DataFrame
    df = pd.DataFrame(list(inventory_items))

    # 重命名列名，使用中文标题
    column_mapping = {'id': '序号',
                      'project_department': '项目部',
                      'blaster': '爆破员',
                      'emulsion_explosive_32mm': '32乳化(公斤)',
                      'powdery_explosive_box_2': '2号粉箱(公斤)',
                      'sticky_explosive': '粘药(公斤)',
                      'electronic_detonator_5m': '5米电子雷管(发)',
                      'electronic_detonator_15m': '15米电子雷管(发)',
                      'inventory_status': '库存状态',
                      'detonating_device_quantity': '起爆具(个)',
                      'detonating_cord_length': '导爆索长度',
                      'date': '日期'
                      }
    df.rename(columns=column_mapping, inplace=True)

    # 处理日期格式
    if '日期' in df.columns:
        df['日期'] = pd.to_datetime(df['日期']).dt.strftime('%Y-%m-%d %H:%M:%S')

    # 处理空值
    df = df.fillna('')

    # ===== 添加按项目部和日期分组汇总的功能 =====
    # 创建汇总数据
    # 复制原始数据，用于汇总计算
    df_for_summary = df.copy()

    # 提取日期的日期部分（不包含时间）用于分组
    df_for_summary['日期'] = pd.to_datetime(df_for_summary['日期']).dt.date

    # 确定需要汇总的数值列
    numeric_columns = [
        '32乳化(公斤)', '2号粉箱(公斤)', '粘药(公斤)',
        '5米电子雷管(发)', '15米电子雷管(发)', '起爆具(个)', '导爆索长度'
    ]

    # 将数值列转换为数值类型
    for col in numeric_columns:
        df_for_summary[col] = pd.to_numeric(
            df_for_summary[col], errors='coerce').fillna(0)

    # 按项目部和日期进行分组汇总
    summary_df = df_for_summary.groupby(['项目部', '日期'], as_index=False)[
        numeric_columns].sum()

    # 在汇总数据中添加标记列
    summary_df['序号'] = '合计'
    summary_df['爆破员'] = ''
    summary_df['库存状态'] = '汇总数据'

    # 计算2号粉箱包装和32乳化包装
    # 2号粉箱：24公斤一箱，3公斤一包
    summary_df['2号粉箱包装'] = summary_df.apply(lambda row:
                                            f"{int(row['2号粉箱(公斤)'] // 24)}箱{int((row['2号粉箱(公斤)'] % 24) // 3)}包"
                                            if pd.notnull(row['2号粉箱(公斤)']) else '', axis=1
                                            )

    # 32乳化：24公斤一箱，6公斤一包
    summary_df['32乳化包装'] = summary_df.apply(lambda row:
                                            f"{int(row['32乳化(公斤)'] // 24)}箱{int((row['32乳化(公斤)'] % 24) // 6)}包"
                                            if pd.notnull(row['32乳化(公斤)']) else '', axis=1
                                            )

    # ===== 按项目部和日期排序，日期近的在上 =====
    # 先将日期列转换为日期类型以便正确排序
    summary_df['日期'] = pd.to_datetime(summary_df['日期'])
    # 按项目部升序、日期降序排序
    summary_df = summary_df.sort_values(
        by=['项目部', '日期'], ascending=[True, False])
    # 将日期转换回字符串格式
    summary_df['日期'] = summary_df['日期'].dt.strftime('%Y-%m-%d')
    # =======================================

    # 确保汇总数据的列顺序与原始数据一致，并添加新列
    original_columns = df.columns.tolist()
    # 在'2号粉箱(公斤)'列后面插入'2号粉箱包装'列
    if '2号粉箱(公斤)' in original_columns:
        index = original_columns.index('2号粉箱(公斤)') + 1
        original_columns.insert(index, '2号粉箱包装')
    # 在'32乳化(公斤)'列后面插入'32乳化包装'列
    if '32乳化(公斤)' in original_columns:
        index = original_columns.index('32乳化(公斤)') + 1
        original_columns.insert(index, '32乳化包装')

    # 确保原始数据也有这两列（填充空字符串）
    for col in ['2号粉箱包装', '32乳化包装']:
        if col not in df.columns:
            df[col] = ''

    # 调整列顺序
    df = df[original_columns]
    summary_df = summary_df[original_columns]

    del summary_df['序号']
    del summary_df['爆破员']
    del summary_df['库存状态']
    # 创建Excel输出流
    excel_io = io.BytesIO()

    # 将数据写入Excel的不同工作表
    with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
        # 写入原始数据到第一个工作表
        df.to_excel(writer, index=False, sheet_name='原始数据')

        # 设置原始数据工作表的列宽
        worksheet_raw = writer.sheets['原始数据']
        # 动态设置列宽，适应新增的列
        for i, col in enumerate(original_columns, 1):
            column_letter = chr(64 + i)  # 64是ASCII码'A'的前一个字符
            worksheet_raw.column_dimensions[column_letter].width = 20

        # 写入汇总数据到第二个工作表
        summary_df.to_excel(writer, index=False, sheet_name='汇总数据')

        # 设置汇总数据工作表的列宽
        worksheet_summary = writer.sheets['汇总数据']
        for i, col in enumerate(original_columns, 1):
            column_letter = chr(64 + i)
            worksheet_summary.column_dimensions[column_letter].width = 15

    excel_io.seek(0)

    # 创建HTTP响应
    response = HttpResponse(
        excel_io.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # 设置文件名
    filename = f"inventory_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# categorycontent------------------------------------------------------------------------------------------------------------------------------
# categorycontent------------------------------------------------------------------------------------------------------------------------------
# categorycontent------------------------------------------------------------------------------------------------------------------------------
# categorycontent------------------------------------------------------------------------------------------------------------------------------


def categorycontent_list(request):

    database = 'categorycontent'
    title = '民爆物品'
    if request.method == "GET":

        data = models.CategoryContent.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at")[:100]

        lst = dframe(data)
        cols = []

        for i in lst:
            cols.append({'age': i})

        return render(request, 'list.html', {"data": data, "cols": cols, "数据库": database, '标题': title})


def categorycontent_create(request):
    title = '民爆物品'

    if request.method == "GET":

        form = modelform.CategoryContentForm()

        return render(request, 'card_form.html', {'form': form, '标题': title})

    if request.method == 'POST':
        form = modelform. CategoryContentForm(request.POST)

        if form.is_valid():
            form.save()

        else:
            form.errors
            return render(request, 'card_form.html', {'form': form, '标题': title})

    return redirect("/home/categorycontent_list")


def categorycontent_delete(request):

    id = request.GET.get('id')
    models.CategoryContent.objects.filter(id=str(id)).delete()
    return redirect("/home/categorycontent_list")


def categorycontent_edit(request):
    title = '民爆物品'
    id = request.GET.get('id')
    row_object = models.CategoryContent.objects.filter(id=str(id)).first()

    if request.method == "GET":

        form = modelform.CategoryContentForm(instance=row_object)

        return render(request, 'card_form.html', {"form": form, "标题": title})

    form = modelform.CategoryContentForm(
        data=request.POST, instance=row_object)

    if form.is_valid():

        form.save()
    else:
        title = '输入错误'
        form.errors
        return render(request, 'card_form.html', {'form': form})
    return redirect("/home/categorycontent_list")


# ---------------------------------------------------------------------------------------


def upload_model(request):
    # 获取指定应用中的所有模型
    app_models = apps.get_app_config('app01').get_models()
    model_names = [model.__name__ for model in app_models]

    if request.method == 'POST':
        model_name = request.POST.get('model_name')
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            # 处理上传的 '.xlsx' 文件
            df = pd.read_excel(uploaded_file)
            df['password'] = df['password'].apply(lambda x: md5(str(x)))
            # 这里可以根据选择的模型表名进行相应的处理
            print(f"选择的模型表: {model_name}")
            print(df.head())
            try:
                with sl.connect('/Users/sunhongchen/lnjx2025/db.sqlite3') as con:
                    df.to_sql(f'app01_{model_name.lower()}',
                              con, index=False, if_exists='append')

                return HttpResponse("文件上传成功！")
            except Exception as e:
                return HttpResponse(f"文件上传失败: {str(e)}")

    return render(request, 'upload.html', {'model_names': model_names})


# --------------------------------------
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
def candidateprofile_add(request):
    title = 'candidateprofile'
    if request.method == "GET":
        form = modelform.CandidateProfileForm()
        return render(request, 'candidateprofile_add.html', {"form": form, "标题": title})

    form = modelform.CandidateProfileForm(
        data=request.POST,
        files=request.FILES
    )

    if form.is_valid():
        # 处理上传的证件照
        img_bytes = None
        if 'photo' in request.FILES:
            photo_file = request.FILES['photo']
            with Image.open(photo_file) as img:
                if img.mode in ('RGBA', 'LA'):
                    img = img.convert('RGB')

                processed_img = resize_photo(cut_photo(img, 1), 1)

                img_io = io.BytesIO()
                processed_img.save(img_io, format='JPEG', quality=90)
                img_bytes = img_io.getvalue()

                form.instance.photo.save(
                    f"{form.cleaned_data['name']}_processed.jpg",
                    ContentFile(img_bytes),
                    save=False
                )

        # 保存表单数据到数据库
        instance = form.save()

        # 创建Excel文件

        wb = Workbook()
        ws = wb.active
        ws.title = "候选人档案"

        # 设置列宽和行高
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35  # 加宽B列适应换行

        # 创建样式对象
        from openpyxl.styles import Font, Alignment
        header_font = Font(size=16)
        cell_alignment = Alignment(vertical='center', wrap_text=True)

        # 写入基础数据
        data_fields = [
            ('姓名', 'name'),
            ('性别', 'gender'),
            ('手机', 'mobile'),
            ('年龄', 'age'),
            ('婚姻状况', 'marital_status'),
            ('学历专业', 'education'),
            ('驾照', 'has_driver_license'),
            ('特长', 'special_skills'),
            ('工作经历', 'work_experience'),
            ('现住所', 'current_address'),
            ('应聘岗位', 'position'),
            ('期望薪资', 'expected_salary')
        ]  # current_address

        for row, (label, field) in enumerate(data_fields, start=1):
            # 写入A列
            cell = ws[f'A{row}']
            cell.value = label
            cell.font = header_font
            cell.alignment = cell_alignment

            # 写入B列
            cell = ws[f'B{row}']
            cell.value = str(form.cleaned_data.get(field, ''))
            cell.font = Font(size=16)
            cell.alignment = cell_alignment

            # 设置行高
            ws.row_dimensions[row].height = 35  # 根据字体大小调整行高

        # 插入照片（如果存在）
        if img_bytes:
            img_temp = io.BytesIO(img_bytes)
            img = ExcelImage(img_temp)
            img.width = 110
            img.height = 160
            ws.add_image(img, 'c1')

        # 保存Excel文件
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)

        # 保存到模型的resume_file字段
        instance.resume_file.save(
            f"{form.cleaned_data['name']}_profile.xlsx",
            ContentFile(excel_io.getvalue()),
            save=True
        )

        return HttpResponse("""
            <div style="
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                margin: 0;
            ">
                <h1 style="font-size: 48px; margin: 0">数据上传成功</h1>
            </div>
        """)

    else:
        title = '输入错误'
        return render(request, 'candidateprofile_add.html', {"form": form, "标题": title})


@资料员
def candidateprofile_delete(request):
    id = request.GET.get('id')
    models.Candidate.objects.filter(id=str(id)).delete()
    return redirect("/home/candidateprofile_list")


def candidateprofile_list(request):
    title = 'candidateprofile'
    if request.method == "GET":
        data = models.Candidate.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at").order_by('-created_at')[:100]
        # 获取模型字段的verbose_name
        model_fields = models.Candidate._meta.fields
        cols = [{'verbose_name': 'id'}] + [{'verbose_name': field.verbose_name}
                                           for field in model_fields if field.name != 'id']

        # 添加操作列
        cols.append({'verbose_name': '操作'})

        return render(request, 'candidateprofile_list.html', {
            "data": data,
            "cols": cols,
            "title": title
        })


# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
def save_image_to_field(img, filename):
    img_io = io.BytesIO()
    if img.mode in ('RGBA', 'LA'):
        img = img.convert('RGB')
    img.save(img_io, format='JPEG')
    return ContentFile(img_io.getvalue(), name=filename)


def process_photo(img):
    img = cut_photo(img, 1)
    img = resize_photo(img, 1)
    img_io = io.BytesIO()
    img.save(img_io, format='JPEG')
    return ContentFile(img_io.getvalue())


# 在文件顶部或适当位置添加方向校正函数
def apply_orientation(img, orientation):
    from PIL import ImageOps
    # 方向值与对应旋转方式
    ORIENTATIONS = {
        1: (0, False),
        2: (0, True),
        3: (180, False),
        4: (180, True),
        5: (90, True),
        6: (270, False),
        7: (270, True),
        8: (90, False)
    }

    degrees, mirror = ORIENTATIONS.get(orientation, (0, False))

    if mirror:
        img = ImageOps.mirror(img)
    if degrees == 90:
        img = img.transpose(Image.ROTATE_90)
    elif degrees == 180:
        img = img.transpose(Image.ROTATE_180)
    elif degrees == 270:
        img = img.transpose(Image.ROTATE_270)

    return img.convert('RGB')  # 确保返回统一格式


def explosivestaff_add(request):
    title = 'tu'
    if request.method == 'POST':
        if 'zip_file' in request.FILES:
            zip_file = request.FILES['zip_file']
            success_count = 0
            errors = []

            try:
                with zipfile.ZipFile(zip_file) as zf:
                    folder_files = defaultdict(dict)
                    for file_info in zf.infolist():
                        # 增强版编码修复（新增多个编码尝试）
                        raw_name = file_info.filename
                        try:
                            # 优先尝试 UTF-8 解码
                            corrected = raw_name.encode(
                                'cp437').decode('utf-8')
                        except UnicodeDecodeError:
                            try:
                                # 次尝试 GB18030 解码
                                corrected = raw_name.encode(
                                    'cp437').decode('gb18030')
                            except:
                                # 最后尝试忽略错误字符
                                corrected = raw_name.encode(
                                    'cp437').decode('utf-8', 'ignore')

                        # 跳过Mac系统文件（新增过滤）
                        if corrected.startswith('._'):
                            continue

                        if not file_info.is_dir() and '/' in corrected:
                            folder, filename = corrected.split('/', 1)
                            folder_files[folder][filename] = file_info

                    for folder, files in folder_files.items():
                        try:
                            required_files = {
                                '人像.jpg': None,
                                '国徽.jpg': None,
                                '证件照.jpg': None,
                                '无犯罪证明.jpg': None,
                                '毕业证.jpg': None
                            }

                            # 读取文件（添加调试输出）

                            for filename in required_files.keys():
                                if filename in files:
                                    try:
                                        file_info = files[filename]
                                        # 添加文件存在性验证
                                        if file_info.file_size == 0:
                                            raise ValueError(
                                                f"文件 {filename} 为空")

                                        with Image.open(io.BytesIO(zf.read(file_info))) as img:
                                            required_files[filename] = img
                                            # 添加图片有效性验证
                                            img.verify()  # 验证图片完整性
                                    except Exception as e:
                                        raise ValueError(
                                            f"文件 {filename} 损坏: {str(e)}")
                                else:
                                    raise ValueError(f"缺少文件: {filename}")

                        # 处理身份证信息

                        # 添加空值检查
                            if not all(required_files.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at")):
                                missing = [
                                    k for k, v in required_files.items() if not v]
                                raise ValueError(f"图片加载失败: {missing}")
                                # 处理身份证信息
                                姓名, 身份证号码 = sfz(required_files['人像.jpg'])
                                print(f"姓名: {姓名}, 身份证号码: {身份证号码}")
                                # 处理各图片并保存
                                front_img = resize_photo(
                                    required_files['人像.jpg'], 3)
                                back_img = resize_photo(
                                    required_files['国徽.jpg'], 3)
                                combined_img = combine_a4_images(
                                    front_img, back_img)

                                # 保存到数据库
                                models.ExplosiveStaff.objects.create(
                                    name=姓名,
                                    id_number=身份证号码,
                                    front_image=save_image_to_field(
                                        front_img, f"{身份证号码}_front.jpg"),
                                    back_image=save_image_to_field(
                                        back_img, f"{身份证号码}_back.jpg"),
                                    combined_image=save_image_to_field(
                                        combined_img, f"{身份证号码}_combined.jpg"),
                                    photo=process_photo(
                                        required_files['证件照.jpg']),
                                    no_crime=save_image_to_field(
                                        required_files['无犯罪证明.jpg'], f"{身份证号码}_no_crime.jpg"),
                                    graduation=save_image_to_field(
                                        required_files['毕业证.jpg'], f"{身份证号码}_graduation.jpg")
                                )
                                success_count += 1

                        except Exception as e:
                            errors.append(f"{folder}: {str(e)}")

                return HttpResponse(f"成功导入 {success_count} 条记录，错误 {len(errors)} 条{errors}")

            except zipfile.BadZipFile:
                return HttpResponse("无效的ZIP文件格式", status=400)

        else:
            # 原有单个文件处理逻辑保持不变
            front_file = request.FILES.get('front')

            back_file = request.FILES.get('back')
            photo = request.FILES.get('photo')
            no_crime = request.FILES.get('no_crime')  # 原错误参数 'photo'
            graduation = request.FILES.get('graduation')  # 原错误参数 'photo'
            mobile = request.POST.get('mobile', '')  # 新增
            bank_card_number = request.POST.get('bank_card_number', '')  # 新增

            with Image.open(front_file) as img1:
                img1 = resize_photo(img1, 3)
                # 新增模式转换
                if img1.mode in ('RGBA', 'LA'):
                    img1 = img1.convert('RGB')
                img_io = io.BytesIO()
                img1.save(img_io, format='JPEG')
                img_人像 = img_io.getvalue()

            with Image.open(back_file) as img2:
                img2 = resize_photo(img2, 3)
                # 新增模式转换
                if img2.mode in ('RGBA', 'LA'):
                    img2 = img2.convert('RGB')
                img_io = io.BytesIO()
                img2.save(img_io, format='JPEG')
                img_国徽 = img_io.getvalue()

            with Image.open(photo) as img3:
                # 添加方向校正
                try:
                    exif = img3.getexif()
                    orientation = exif.get(274)  # 274是Exif方向标签
                    if orientation:
                        img3 = apply_orientation(img3, orientation)
                except Exception as e:
                    print(f"方向校正失败: {str(e)}")

                if img3.mode in ('RGBA', 'LA'):
                    img3 = img3.convert('RGB')  # 移除Alpha通道
                img_io = io.BytesIO()
                img3 = resize_photo(cut_photo(img3, 1), 1)
                img3.save(img_io, format='JPEG')
                img_bytes = img_io.getvalue()
                rotated_io = io.BytesIO()
                排版(img3, rotated_io)
                rotated_bytes = rotated_io.getvalue()

            # 处理无犯罪证明（带存在性检查）
            img_数据4 = None
            if no_crime:  # 添加文件存在判断
                with Image.open(no_crime) as img4:
                    img4 = resize_photo(img4, 4)
                    img_io = io.BytesIO()
                    img4.save(img_io, format='JPEG')
                    img_数据4 = img_io.getvalue()

            # 处理毕业证（带存在性检查）
            img_数据5 = None
            if graduation:  # 添加文件存在判断
                with Image.open(graduation) as img5:
                    img5 = resize_photo(img5, 4)
                    img_io = io.BytesIO()
                    img5.save(img_io, format='JPEG')
                    img_数据5 = img_io.getvalue()

            # 合并图片逻辑保持不变
            combined_img = combine_a4_images(img1, img2)
            img_io = io.BytesIO()
            combined_img.save(img_io, format='JPEG', quality=90)
            combined_bytes = img_io.getvalue()

            try:
                姓名, 身份证号码 = sfz(front_file)
            except:
                姓名 = "未识别"
                身份证号码 = "888888888888888888"  # 18个8

            # 修改模型引用为ExplosiveStaff
            models.ExplosiveStaff.objects.create(
                name=姓名,
                id_number=身份证号码,
                mobile=mobile,  # 新增字段
                bank_card_number=bank_card_number,
                front_image=ContentFile(img_人像, name=f"{身份证号码}_front.jpg"),
                back_image=ContentFile(img_国徽, name=f"{身份证号码}_back.jpg"),
                combined_image=ContentFile(
                    combined_bytes, name=f"{身份证号码}_combined.jpg"),

                photo=ContentFile(img_bytes, name=f"{身份证号码}_photo.jpg"),
                typeset_photo=ContentFile(
                    rotated_bytes, name=f"{身份证号码}_typeset.jpg"),
                # ... existing code ...
                no_crime=ContentFile(
                    img_数据4, name=f"{身份证号码}_no_crime.jpg") if img_数据4 else None,
                graduation=ContentFile(
                    img_数据5, name=f"{身份证号码}_graduation.jpg") if img_数据5 else None,
                # ... existing code ...

            )

            return HttpResponse("""
                <div style="
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100vh;
                    margin: 0;
                ">
                    <h1 style="font-size: 48px; margin: 0">数据上传成功</h1>
                </div>
            """)

    else:
        model_names = models.ExplosiveStaff.objects.values_list(
            'id_number', flat=True).distinct()
        return render(request, 'explosivestaff_add.html', {
            'model_names': model_names,
            'zip_upload': True
        })


def explosivestaff_delete(request):
    id = request.GET.get('id')
    models.ExplosiveStaff.objects.filter(id=str(id)).delete()
    return redirect("/home/explosivestaff_list")  # 修改重定向路径


def explosivestaff_list(request):
    title = 'explosivestaff'
    if request.method == "GET":
        model_fields = models.ExplosiveStaff._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields if field.attname not in ('id', 'location')]
        cols.append({'verbose_name': '操作'})
        data = models.ExplosiveStaff.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at").order_by(
            '-created_at')[:100]
        return render(request, 'explosivestaff_list.html', {
            "data": data,
            "cols": cols,
            "title": title,
            "export_url": "/home/explosivestaff_export_zip",
            "export_xlsx_url": "/home/explosivestaff_export_xlsx"  # 新增XLSX导出参数
        })
# 新增Excel导出函数


def explosivestaff_export_xlsx(request):
    from openpyxl import Workbook
    from django.http import HttpResponse
    import io
    from datetime import datetime

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "爆破员数据"

    # 设置标题行
    headers = ['姓名', '身份证号码', '手机号', '银行卡号']
    ws.append(headers)

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

    # 获取数据
    staffs = models.ExplosiveStaff.objects.all().values_list(
        'name', 'id_number', 'mobile', 'bank_card_number'
    )

    # 填充数据
    for staff in staffs:
        ws.append([
            staff[0],  # 姓名
            staff[1],  # 身份证号码
            staff[2] or '',  # 手机号
            staff[3] or ''   # 银行卡号
        ])

    # 创建HTTP响应
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    response = HttpResponse(
        excel_io.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"explosivestaff_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
# 新增的导出函数


def explosivestaff_export_zip(request):

    # 创建内存文件
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # 获取所有记录
        staffs = models.ExplosiveStaff.objects.all()

        for staff in staffs:
            # 添加所有图片字段到ZIP
            fields = ['front_image', 'back_image', 'combined_image',
                      'photo', 'typeset_photo', 'no_crime', 'graduation']

            for field in fields:
                file = getattr(staff, field)
                if file and file.storage.exists(file.name):
                    zipf.writestr(
                        f"{staff.name}_{staff.id_number}/{field}.jpg",
                        file.read()
                    )

    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="explosivestaff_images.zip"'
    return response


# weighingrecords333333333333333333333333#weighingrecords333333333333333333333333#weighingrecords333333333333333333333333
# weighingrecords333333333333333333333333#weighingrecords333333333333333333333333#weighingrecords333333333333333333333333
# weighingrecords333333333333333333333333#weighingrecords333333333333333333333333#weighingrecords333333333333333333333333
# weighingrecords333333333333333333333333#weighingrecords333333333333333333333333#weighingrecords333333333333333333333333


def weighingrecord_add(request):
    title = 'weighingrecord'
    if request.method == "GET":

        form = modelform.WeighingRecordForm()

        return render(request, 'card_form.html', {'form': form, '标题': title})

    if request.method == 'POST':
        form = modelform.WeighingRecordForm(
            request.POST, request.FILES)  # <-- 添加request.FILES

        if form.is_valid():

            form.save()

        else:

            form.errors
            return render(request, 'card_form.html', {'form': form, '标题': title})

    return redirect("/home/weighingrecord_list")


@资料员
def weighingrecord_delete(request):
    id = request.GET.get('id')
    models.WeighingRecord.objects.filter(id=str(id)).delete()
    return redirect("/home/weighingrecord_list")


def weighingrecord_list(request):
    title = 'weighingrecord'
    if request.method == "GET":
        data = models.WeighingRecord.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at").order_by(
            'weight_number', '-created_at')
        # 获取模型字段的verbose_name
        model_fields = models.WeighingRecord._meta.fields
        cols = [{'verbose_name': 'id'}] + [{'verbose_name': field.verbose_name}
                                           for field in model_fields if field.name != 'id']

        # 添加操作列
        cols.append({'verbose_name': '操作'})

        return render(request, 'weighingrecord_list.html', {
            "data": data,
            "cols": cols,
            "title": title
        })


# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################

@csrf_exempt

def idcard_batch_upload(request):
    """接收一张正反面拼图 → OCR分割 → 入库"""
    from django.views.decorators.csrf import csrf_exempt
    import tempfile, os, cv2, math
    import numpy as np
    from PIL import Image as PILImage
    from microwink import SegModel
    import onnxruntime as ort
    from rapidocr_onnxruntime.ch_ppocr_det.text_detect import TextDetector
    from django.core.files.uploadedfile import InMemoryUploadedFile

    if request.method != 'POST':
        return redirect("/home/idcard_list/")

    uploaded = request.FILES.get('image')
    if not uploaded:
        return redirect("/home/idcard_list/")

    # 模型路径
    _model_dir = os.path.expanduser("~/.hermes/models")
    if not os.path.exists(os.path.join(_model_dir, "seg_model.onnx")):
        return redirect("/home/idcard_list/")

    SEG_PATH = os.path.join(_model_dir, "seg_model.onnx")
    DET_PATH = os.path.join(_model_dir, "det.onnx")
    REC_PATH = os.path.join(_model_dir, "rec.onnx")
    CLS_PATH = os.path.join(_model_dir, "cls.onnx")
    for p in [SEG_PATH, DET_PATH, REC_PATH, CLS_PATH]:
        if not os.path.exists(p):
            return redirect("/home/idcard_list/")

    temp_dir = tempfile.mkdtemp()
    try:
        # 保存上传图片
        img_path = os.path.join(temp_dir, uploaded.name)
        with open(img_path, 'wb') as f:
            for chunk in uploaded.chunks():
                f.write(chunk)

        img_cv = cv2.imread(img_path)
        if img_cv is None:
            return redirect("/home/idcard_list/")

        # ====== split_smart ======
        def find_balance_split_line(mask, thresh=0.5, step=2, min_ratio=0.15):
            h = mask.shape[0]
            binary = (mask > thresh).astype(np.uint8)
            total_px = np.sum(binary)
            min_side_px = total_px * min_ratio
            valid = []
            for y in range(0, h, step):
                up = np.sum(binary[:y, :])
                down = np.sum(binary[y:, :])
                if up >= min_side_px and down >= min_side_px:
                    valid.append((y, np.sum(binary[y, :])))
            return min(valid, key=lambda x: x[1])[0] if valid else h // 2

        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        _, binary = cv2.threshold(gray, 240, 1, cv2.THRESH_BINARY_INV)
        total_px = np.sum(binary)
        h = img_cv.shape[0]
        seam_y = int(h * 0.2) + int(np.argmin(np.sum(binary, axis=1)[int(h*0.2):int(h*0.8)]))
        up_px = np.sum(binary[:seam_y, :])
        down_px = np.sum(binary[seam_y:, :])
        min_side = total_px * 0.15
        if up_px >= min_side and down_px >= min_side:
            split_y = seam_y
        else:
            split_y = find_balance_split_line(binary)
        top_cv = img_cv[:split_y]
        bottom_cv = img_cv[split_y:]

        # ====== microwink ======
        seg = SegModel.from_path(SEG_PATH)
        def microwink_crop(img, tw=856, th=540):
            cards = seg.apply(PILImage.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB)))
            if not cards: return None
            mask = (cards[0].mask > 0.5).astype(np.uint8) * 255
            cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if not cnts: return None
            cnt = max(cnts, key=cv2.contourArea)
            approx = cv2.approxPolyDP(cnt, 0.02*cv2.arcLength(cnt, True), True).reshape(4, 2)
            pts = sorted(approx, key=lambda x: (x[1], x[0]))
            if pts[0][0] > pts[1][0]: pts[0], pts[1] = pts[1], pts[0]
            if pts[2][0] > pts[3][0]: pts[2], pts[3] = pts[3], pts[2]
            src = np.array([pts[0], pts[1], pts[3], pts[2]], dtype=np.float32)
            dst = np.array([[0,0],[tw,0],[tw,th],[0,th]], dtype=np.float32)
            return cv2.warpPerspective(img, cv2.getPerspectiveTransform(src, dst),
                                        (tw, th), borderMode=cv2.BORDER_REPLICATE)

        # ====== ClsModel ======
        class ClsModel:
            def __init__(self, path):
                self.s = ort.InferenceSession(path, providers=["CPUExecutionProvider"])
                self.imgH, self.imgW, self.thresh = 48, 192, 0.9
            def resize(self, img):
                h, w = img.shape[:2]
                rw = self.imgW if math.ceil(self.imgH*w/h) > self.imgW else int(math.ceil(self.imgH*w/h))
                n = cv2.resize(img, (rw, self.imgH)).astype("float32").transpose(2,0,1)/255
                n = (n-0.5)/0.5; p = np.zeros((3,self.imgH,self.imgW), dtype=np.float32)
                p[:,:,:rw] = n; return p
            def predict(self, crop):
                prob = self.s.run(None, {"x": self.resize(crop)[np.newaxis,:].astype(np.float32)})[0]
                return float(prob[0][0]), float(prob[0][1])
            def orient_page(self, img, det):
                dt_boxes, _ = det(img)
                if dt_boxes is None or len(dt_boxes) == 0: return img, False
                flipped = 0
                for box in dt_boxes:
                    w = int(max(np.linalg.norm(box[0]-box[1]), np.linalg.norm(box[2]-box[3])))
                    hb = int(max(np.linalg.norm(box[0]-box[3]), np.linalg.norm(box[1]-box[2])))
                    M = cv2.getPerspectiveTransform(box.astype(np.float32),
                        np.array([[0,0],[w,0],[w,hb],[0,hb]], dtype=np.float32))
                    crop = cv2.warpPerspective(img, M, (w,hb), borderMode=cv2.BORDER_REPLICATE)
                    p0, p1 = self.predict(crop)
                    if p1 >= self.thresh: flipped += 1
                if flipped > len(dt_boxes) // 2:
                    return cv2.rotate(img, cv2.ROTATE_180), True
                return img, False

        det = TextDetector({"model_path": DET_PATH, "use_cuda": False, "limit_side_len": 736, "limit_type": "min",
            "std":[0.5]*3,"mean":[0.5]*3,"thresh":0.3,"box_thresh":0.5,"use_dilation":True,"score_mode":"fast"})
        cls_model = ClsModel(CLS_PATH)

        # 处理上下两半
        halves = {}
        for side, half in [("上", top_cv), ("下", bottom_cv)]:
            cropped = microwink_crop(half)
            if cropped is None:
                return redirect("/home/idcard_list/")
            cropped, _ = cls_model.orient_page(cropped, det)
            halves[side] = cropped

        # ====== OCR + 内容判断 ======
        class RecModel:
            def __init__(self, path):
                self.s = ort.InferenceSession(path, providers=["CPUExecutionProvider"])
                meta = self.s.get_modelmeta().custom_metadata_map
                self.chars = ["blank"] + meta["character"].splitlines() + [" "]
            def __call__(self, img_list):
                if not img_list: return []
                ws = [i.shape[1]/i.shape[0] for i in img_list]
                idx = np.argsort(ws)
                res = [("", 0.0)] * len(img_list)
                for b in range(0, len(img_list), 6):
                    bi = idx[b:min(len(img_list), b+6)]
                    mw = max(max(ws[i] for i in bi), 320/48)
                    iw = int(48*mw)
                    batch = []
                    for i in bi:
                        rw = min(iw, int(math.ceil(48*img_list[i].shape[1]/img_list[i].shape[0])))
                        n = cv2.resize(img_list[i], (rw, 48)).astype("float32").transpose(2,0,1)/255
                        n = (n-0.5)/0.5; p = np.zeros((3,48,iw), dtype=np.float32)
                        p[:,:,:rw] = n; batch.append(p[np.newaxis,:])
                    preds = self.s.run(None, {"x": np.concatenate(batch).astype(np.float32)})[0]
                    for rno in range(preds.shape[0]):
                        ix, pr = preds[rno].argmax(1), preds[rno].max(1)
                        cs, cn, pv = [], [], -1
                        for j, cid in enumerate(ix):
                            if cid != 0 and cid != pv:
                                cs.append(self.chars[cid] if cid < len(self.chars) else "")
                                cn.append(float(pr[j]))
                            pv = int(cid)
                        res[bi[rno]] = ("".join(cs), np.mean(cn) if cn else 0.0)
                return res

        rec = RecModel(REC_PATH)

        ocr_texts = {}
        for sn in ["上", "下"]:
            dt_boxes, _ = det(halves[sn])
            lines = []
            if dt_boxes is not None and len(dt_boxes) > 0:
                sboxes = sorted(enumerate(dt_boxes), key=lambda x: (x[1][0][1], x[1][0][0]))
                crops = []
                for i, box in sboxes:
                    w = int(max(np.linalg.norm(box[0]-box[1]), np.linalg.norm(box[2]-box[3])))
                    hb = int(max(np.linalg.norm(box[0]-box[3]), np.linalg.norm(box[1]-box[2])))
                    M = cv2.getPerspectiveTransform(box.astype(np.float32),
                        np.array([[0,0],[w,0],[w,hb],[0,hb]], dtype=np.float32))
                    crop = cv2.warpPerspective(halves[sn], M, (w,hb), borderMode=cv2.BORDER_REPLICATE)
                    prob = cls_model.s.run(None, {"x": cls_model.resize(crop)[np.newaxis,:].astype(np.float32)})[0]
                    if prob[0][1] >= cls_model.thresh:
                        crop = cv2.rotate(crop, cv2.ROTATE_180)
                    crops.append(crop)
                rec_res = rec(crops)
                for (i, box), (t, s) in zip(sboxes, rec_res):
                    if s >= 0.5 and t.strip():
                        lines.append(t)
            ocr_texts[sn] = " ".join(lines)

        # 判断人像面/国徽面
        renyi_kw = ["姓名", "公民身份号码"]
        def has_kw(text): return any(kw in text for kw in renyi_kw)
        if has_kw(ocr_texts["上"]):
            renyi_side, guohui_side = "上", "下"
        elif has_kw(ocr_texts["下"]):
            renyi_side, guohui_side = "下", "上"
        else:
            renyi_side, guohui_side = "上", "下"

        # 保存临时图片 + 调用 idcard_add 逻辑
        front_path = os.path.join(temp_dir, "人像面.jpg")
        back_path = os.path.join(temp_dir, "国徽面.jpg")
        PILImage.fromarray(cv2.cvtColor(halves[renyi_side], cv2.COLOR_BGR2RGB)).save(front_path, format='JPEG', quality=95)
        PILImage.fromarray(cv2.cvtColor(halves[guohui_side], cv2.COLOR_BGR2RGB)).save(back_path, format='JPEG', quality=95)

        # 调用 sfz 识别姓名和身份证号
        name, id_number = sfz(front_path)

        # 打开图片, resize, 合成
        with PILImage.open(front_path) as i1, PILImage.open(back_path) as i2:
            i1 = resize_photo(i1, 3)
            i2 = resize_photo(i2, 3)
            if i1.mode == 'RGBA': i1 = i1.convert('RGB')
            if i2.mode == 'RGBA': i2 = i2.convert('RGB')

            img_人像_io = io.BytesIO()
            i1.save(img_人像_io, format='JPEG')
            img_人像_bytes = img_人像_io.getvalue()

            img_国徽_io = io.BytesIO()
            i2.save(img_国徽_io, format='JPEG')
            img_国徽_bytes = img_国徽_io.getvalue()

            combined = combine_a4_images(i1, i2)
            combined_io = io.BytesIO()
            combined.save(combined_io, format='JPEG', quality=90)
            combined_bytes = combined_io.getvalue()

        # 入库
        models.IDCard.objects.create(
            name=name,
            id_number=id_number,
            front_image=ContentFile(img_人像_bytes, name=f"{id_number}.jpg"),
            back_image=ContentFile(img_国徽_bytes, name=f"{id_number}_rotated.jpg"),
            combined_image=ContentFile(combined_bytes, name=f"{id_number}_双面.jpg"),
        )

        return redirect("/home/idcard_list/")

    except Exception as e:
        import traceback
        traceback.print_exc()
        return redirect("/home/idcard_list/")
    finally:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)


def test_upload(request):
    from django import forms
    class TestUploadForm(forms.Form):
        file = forms.ImageField(label='选择图片')
    if request.method == 'POST':
        form = TestUploadForm(request.POST, request.FILES)
        if form.is_valid():
            return HttpResponse('上传成功')
    return render(request, 'upload_base.html', {'form': TestUploadForm(), 'title': '测试上传'})

