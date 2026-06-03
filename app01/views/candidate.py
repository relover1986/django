from django.shortcuts import render, HttpResponse, redirect
import time
import torch
import torch.nn as nn
import json
import tempfile
import cv2
from rapidocr_onnxruntime import RapidOCR
from django.http import JsonResponse
from app01 import models
from app01.models import UploadedZhaopian
from app01 import modelform
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_protect, csrf_exempt
import zipfile
import io
from collections import defaultdict
from app01.services import export_service
from app01.jiami import md5
from app01.func import *
from app01.photo import *
from django_filters.views import FilterView
from django.views.generic import ListView
import pandas as pd
from openpyxl import load_workbook
import os
from docx import Document
import sqlite3 as sl
from django.shortcuts import render
from django.apps import apps
from app01.models import ExplosiveInventoryItem
from app01.modelform import ExplosiveInventoryItemForm
from PIL import Image
from django.core.files.base import ContentFile
from app01.services.utility_service import save_image_to_field, process_photo, apply_orientation
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import base64
from django.core.validators import MinLengthValidator, MaxLengthValidator, RegexValidator
from django.conf import settings
import uuid
import sys
sys.path.insert(0, "/root/django")
import utils
sys.path.insert(0, "/root/django")

from django.http import HttpResponse
from datetime import datetime


from io import BytesIO

from aip import AipBodyAnalysis
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework import status
# --------------------------------------
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
def candidateprofile_add(request):
    title = 'candidateprofile'
    if request.method == "GET":
        form = modelform.CandidateProfileForm()
        return render(request, 'candidateprofile_add.html', {"form": form, "标题": title})

    form = modelform.CandidateProfileForm(
        data=request.POST,
        files=request.FILES
    )

    if form.is_valid():
        # 处理上传的证件照
        img_bytes = None
        if 'photo' in request.FILES:
            photo_file = request.FILES['photo']
            with Image.open(photo_file) as img:
                if img.mode in ('RGBA', 'LA'):
                    img = img.convert('RGB')

                processed_img = resize_photo(cut_photo(img, 1), 1)

                img_io = io.BytesIO()
                processed_img.save(img_io, format='JPEG', quality=90)
                img_bytes = img_io.getvalue()

                form.instance.photo.save(
                    f"{form.cleaned_data['name']}_processed.jpg",
                    ContentFile(img_bytes),
                    save=False
                )

        # 保存表单数据到数据库
        instance = form.save()

        # 创建Excel文件

        wb = Workbook()
        ws = wb.active
        ws.title = "候选人档案"

        # 设置列宽和行高
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35  # 加宽B列适应换行

        # 创建样式对象
        from openpyxl.styles import Font, Alignment
        header_font = Font(size=16)
        cell_alignment = Alignment(vertical='center', wrap_text=True)

        # 写入基础数据
        data_fields = [
            ('姓名', 'name'),
            ('性别', 'gender'),
            ('手机', 'mobile'),
            ('年龄', 'age'),
            ('婚姻状况', 'marital_status'),
            ('学历专业', 'education'),
            ('驾照', 'has_driver_license'),
            ('特长', 'special_skills'),
            ('工作经历', 'work_experience'),
            ('现住所', 'current_address'),
            ('应聘岗位', 'position'),
            ('期望薪资', 'expected_salary')
        ]  # current_address

        for row, (label, field) in enumerate(data_fields, start=1):
            # 写入A列
            cell = ws[f'A{row}']
            cell.value = label
            cell.font = header_font
            cell.alignment = cell_alignment

            # 写入B列
            cell = ws[f'B{row}']
            cell.value = str(form.cleaned_data.get(field, ''))
            cell.font = Font(size=16)
            cell.alignment = cell_alignment

            # 设置行高
            ws.row_dimensions[row].height = 35  # 根据字体大小调整行高

        # 插入照片（如果存在）
        if img_bytes:
            img_temp = io.BytesIO(img_bytes)
            img = ExcelImage(img_temp)
            img.width = 110
            img.height = 160
            ws.add_image(img, 'c1')

        # 保存Excel文件
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)

        # 保存到模型的resume_file字段
        instance.resume_file.save(
            f"{form.cleaned_data['name']}_profile.xlsx",
            ContentFile(excel_io.getvalue()),
            save=True
        )

        return HttpResponse("""
            <div style="
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                margin: 0;
            ">
                <h1 style="font-size: 48px; margin: 0">数据上传成功</h1>
            </div>
        """)

    else:
        title = '输入错误'
        return render(request, 'candidateprofile_add.html', {"form": form, "标题": title})


@require_role("爆破工程技术人员", "资料员")
def candidateprofile_delete(request):
    id = request.GET.get('id')
    models.Candidate.objects.filter(id=str(id)).delete()
    return redirect("/home/candidateprofile_list")


def candidateprofile_list(request):
    title = 'candidateprofile'
    if request.method == "GET":
        data = models.Candidate.objects.values("name", "photo", "rotated_photo", "blue_background", "red_background", "white_background", "uploaded_at").order_by('-created_at')[:100]
        # 获取模型字段的verbose_name
        model_fields = models.Candidate._meta.fields
        cols = [{'verbose_name': 'id'}] + [{'verbose_name': field.verbose_name}
                                           for field in model_fields if field.name != 'id']

        # 添加操作列
        cols.append({'verbose_name': '操作'})

        return render(request, 'candidateprofile_list.html', {
            "data": data,
            "cols": cols,
            "title": title
        })


