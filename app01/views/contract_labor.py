from django.shortcuts import render, HttpResponse, redirect
import time
import torch
import torch.nn as nn
import json
import tempfile
import cv2
from rapidocr_onnxruntime import RapidOCR
from django.http import JsonResponse
from app01 import models
from app01.models import UploadedZhaopian
from app01 import modelform
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_protect, csrf_exempt
import zipfile
import io
from app01.services import export_service
from collections import defaultdict
from app01.jiami import md5
from app01.func import *
from app01.photo import *
from django_filters.views import FilterView
from django.views.generic import ListView
import pandas as pd
from openpyxl import load_workbook
import os
from docx import Document
import sqlite3 as sl
from django.shortcuts import render
from django.apps import apps
from app01.models import ExplosiveInventoryItem
from app01.modelform import ExplosiveInventoryItemForm
from PIL import Image
import io
import os
from django.core.files.base import ContentFile
from openpyxl import Workbook
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage
import base64
from django.core.validators import MinLengthValidator, MaxLengthValidator, RegexValidator
from django.conf import settings
import uuid
import json
import sys
sys.path.insert(0, "/root/django")
import utils
import sys
sys.path.insert(0, "/root/django")
import utils

from openpyxl import Workbook
from django.http import HttpResponse
import io
from datetime import datetime


import zipfile
from io import BytesIO

from aip import AipBodyAnalysis
import base64
import io
from PIL import Image
import os
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework import status
# 百度API配置
APP_ID = '118049497'
API_KEY = 'AmK3oZpZhns9jAm2rJgzRyLq'
SECRET_KEY = 'LbbCCzQyv1FlytQxBHstZ5Yt5i4B7pMw'
client = AipBodyAnalysis(APP_ID, API_KEY, SECRET_KEY)


def contractlabor_add(request):
    title = 'tu'
    if request.method == 'POST':
        # 删除原来的 model_name 获取逻辑
        files = request.FILES.getlist('file')
        for file in files:
            time.sleep(0.1)
            if file.name.endswith('.xlsx'):
                wb = load_workbook(file)
            elif file.name.endswith('.docx'):
                document = Document(file)

        ws = wb.worksheets[0]
        row_No = ws.max_row+1
        col_No = ws.max_column+1
        for table_row in range(2, row_No):
            for file in files:
                if file.name.endswith('.xlsx'):
                    wb = load_workbook(file)
                elif file.name.endswith('.docx'):
                    document = Document(file)

            for table_col in range(1, col_No):
                合同(str(ws.cell(1, table_col).value), str(
                    ws.cell(table_row, table_col).value), document)

            姓名 = str(ws.cell(table_row, 3).value)
            img_io = io.BytesIO()

            document.save(img_io)
            document_bytes = img_io.getvalue()  # 直接获取字节数据

            # 创建模型实例
            models.ContractLabor.objects.create(
                name=str(ws.cell(table_row, 3).value),
                id_number=str(ws.cell(table_row, 7).value),
                contract_file=ContentFile(document_bytes, name=f"{姓名}.docx"),

            )

        return redirect("/home/contractlabor_list")

    else:
        model_names = models.ContractLabor.objects.values_list(
            'name', flat=True).distinct()
        return render(request, 'contractlabor_add.html', {'model_names': model_names})


def contractlabor_delete(request):

    id = request.GET.get('id')
    models.ContractLabor.objects.filter(id=str(id)).delete()
    return redirect("/home/contractlabor_list")


def contractlabor_list(request):
    title = 'contractlabor'
    if request.method == "GET":
        data = models.ContractLabor.objects.values("name", "id_number", "contract_file", "created_at").order_by('-id')[:100]
        model_fields = models.ContractLabor._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields if field.attname not in ('id', 'location')]

        # 添加操作列
        cols.append({'verbose_name': '操作'})
        return render(request, 'contractlabor_list.html', {
            "data": data,
            "cols": cols,
            "title": title,
            "export_url": "/home/contractlabor_export_zip"  # 新增导出参数
        })

# 新增合同导出函数


@require_role("爆破工程技术人员", "资料员")
def contractlabor_export_zip(request):
    zip_buffer = export_service.contractlabor_export_zip()
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="labor_contracts.zip"'
    return response

