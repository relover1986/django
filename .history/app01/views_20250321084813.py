from django.shortcuts import render, HttpResponse, redirect
from app01 import models
from app01 import modelform
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_protect
import zipfile
import io
from collections import defaultdict
from app01.jiami import md5
from .func import *
from .photo import *
from django_filters.views import FilterView
from django.views.generic import ListView
import pandas as pd
from openpyxl import load_workbook
import os
from docx import Document
import sqlite3 as sl
from django.shortcuts import render
from django.apps import apps
from app01.models import ExplosiveInventoryItem
from app01.modelform import ExplosiveInventoryItemForm
from PIL import Image
import io
import os  # 需要添加导入
from django.core.files.base import ContentFile  # 新增导入
from openpyxl import Workbook
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage# 添加图片类导入
from django.core.validators import MinLengthValidator, MaxLengthValidator, RegexValidator

import zipfile
from io import BytesIO

from aip import AipBodyAnalysis
import base64
import io
from PIL import Image
import os

# 百度API配置
APP_ID = '118049497'
API_KEY = 'AmK3oZpZhns9jAm2rJgzRyLq'
SECRET_KEY = 'LbbCCzQyv1FlytQxBHstZ5Yt5i4B7pMw'
client = AipBodyAnalysis(APP_ID, API_KEY, SECRET_KEY)














def home(request):
    return render(request, "home.html")

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def contractlabor_add(request):
    title = 'tu'
    if request.method == 'POST':
        # 删除原来的 model_name 获取逻辑
        files = request.FILES.getlist('file')
        
            

        for file in files:
            if file.name.endswith('.xlsx'):
                wb= load_workbook(file)
            elif file.name.endswith('.docx'): 
                document = Document(file)
            
        ws = wb.worksheets[0]
        row_No=ws.max_row+1
        col_No=ws.max_column+1
        for table_row in range(2, row_No):
            for file in files:
                if file.name.endswith('.xlsx'):
                    wb= load_workbook(file)
                elif file.name.endswith('.docx'): 
                    document = Document(file)
                    
            for table_col in range(1, col_No):
                合同(str(ws.cell(1, table_col).value), str(ws.cell(table_row, table_col).value),document)
        
    
            姓名=str(ws.cell(table_row, 3).value)
            img_io = io.BytesIO()
            

            document.save(img_io)
            document_bytes = img_io.getvalue()  # 直接获取字节数据



            # 创建模型实例
            models.ContractLabor.objects.create(
                name=str(ws.cell(table_row, 3).value),
                id_number = str(ws.cell(table_row, 7).value),
                contract_file=ContentFile(document_bytes, name=f"{姓名}.docx"),

            )

 
        return redirect("/home/contractlabor_list")

    else:
        model_names = models.ContractLabor.objects.values_list(
            'name', flat=True).distinct()
        return render(request, 'contractlabor_add.html', {'model_names': model_names})


def contractlabor_delete(request):

    id = request.GET.get('id')
    models.ContractLabor.objects.filter(id=str(id)).delete()
    return redirect("/home/contractlabor_list")


def contractlabor_list(request):
    title = 'contractlabor'
    if request.method == "GET":
        data = models.ContractLabor.objects.values().order_by('-id')[:100]
        model_fields = models.ContractLabor._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields ]
        
        # 添加操作列
        cols.append({'verbose_name': '操作'})
        return render(request, 'contractlabor_list.html', {
            "data": data,
            "cols": cols,
            "title": title,
            "export_url": "/home/contractlabor_export_zip"  # 新增导出参数
        })

# 新增合同导出函数
@资料员
def contractlabor_export_zip(request):
    import zipfile
    from io import BytesIO
    
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # 获取所有合同工
        laborers = models.ContractLabor.objects.all()
        
        for laborer in laborers:
            # 获取合同文件字段
            contract_file = laborer.contract_file
            if contract_file and contract_file.storage.exists(contract_file.name):
                # 使用身份证号+姓名作为目录名
                zipf.writestr(
                    f"{laborer.id_number}_{laborer.name}/劳动合同.docx",
                    contract_file.read()
                )
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="labor_contracts.zip"'
    return response


#photo++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def photo_add(request):
    title = 'tu'
    if request.method == 'POST':
        # 删除原来的 model_name 获取逻辑
        files = request.FILES.getlist('file')
        
            

        for file in files:
            
            with Image.open(file) as img:  
                
                if img.mode in ('RGBA', 'LA'):
                    img = img.convert('RGB')  # 移除Alpha通道
                    
                
                
                if len(files)==1:
      
                    filename = request.POST['model_name']
                    if len(filename)==0:
                        filename = os.path.splitext(file.name)[0]

                    
                else:
                    filename = os.path.splitext(file.name)[0]
                    
                
                
                img_io = io.BytesIO()
                
                img=resize_photo(cut_photo(img,1),1)
                img.save(img_io, format='JPEG')
                img_bytes = img_io.getvalue()  # 直接获取字节数据
                
                



                result = client.bodySeg(img_bytes)
                
                if 'foreground' in result:
                    # 转换前景图
                    foreground = Image.open(io.BytesIO(base64.b64decode(result['foreground'])))
                    
                    # 一寸照片标准尺寸
                    output_size = (295, 413)
                    foreground.thumbnail(output_size)
                    
                    # 创建三种背景色
                    background_colors = {
                        '蓝底': (67, 142, 219),
                        '红底': (255, 0, 0),
                        '白底': (255, 255, 255)
                    }
                    
                    # 存储排版后的背景图
                    bg_files = {
                        'blue': None,
                        'red': None,
                        'white': None
                    }

                    # 生成并排版所有背景图
                    for name, color in background_colors.items():
                        # 生成背景
                        background = Image.new('RGB', output_size, color)
                        x = (output_size[0] - foreground.width) // 2
                        y = (output_size[1] - foreground.height) // 2
                        background.paste(foreground, (x, y), foreground)
                        
                        # 对背景图进行排版
                        rotated_bg_io = io.BytesIO()
                        排版(background, rotated_bg_io)
                        rotated_bg_bytes = rotated_bg_io.getvalue()
                        
                        # 根据背景类型存储
                        if name == '蓝底':
                            bg_files['blue'] = ContentFile(rotated_bg_bytes, name=f"{filename}_blue.jpg")
                        elif name == '红底':
                            bg_files['red'] = ContentFile(rotated_bg_bytes, name=f"{filename}_red.jpg")
                        elif name == '白底':
                            bg_files['white'] = ContentFile(rotated_bg_bytes, name=f"{filename}_white.jpg")

                    # 对原始图进行排版
                    rotated_io = io.BytesIO()
                    排版(img, rotated_io)
                    rotated_bytes = rotated_io.getvalue()

                    # 创建模型实例并保存所有图片
                    models.UploadedZhaopian.objects.create(
                        name=filename,
                        photo=ContentFile(img_bytes, name=f"{filename}.jpg"),
                        rotated_photo=ContentFile(rotated_bytes, name=f"{filename}_rotated.jpg"),
                        blue_background=bg_files['blue'],
                        red_background=bg_files['red'],
                        white_background=bg_files['white']
                    )


        return redirect("/home/photo_list")

    else:
        model_names = models.UploadedZhaopian.objects.values_list(
            'name', flat=True).distinct()
        return render(request, 'photo_add.html', {'model_names': model_names})


def photo_delete(request):

    id = request.GET.get('id')
    models.UploadedZhaopian.objects.filter(id=str(id)).delete()
    return redirect("/home/photo_list")


def photo_list(request):

    title = 'photo'
    if request.method == "GET":
        model_fields = models.UploadedZhaopian._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields]
        cols.append({'verbose_name': '操作'})
        data = models.UploadedZhaopian.objects.values().order_by('-uploaded_at')[:100]
        return render(request, 'photo_list.html', {
            "data": data,
            "cols": cols,
            "title": title,
            "export_url": "/home/photo_export_zip"  # 新增导出URL参数
        })

@资料员
def photo_export_zip(request):
    import zipfile
    from io import BytesIO
    
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # 获取所有照片记录
        photos = models.UploadedZhaopian.objects.all()
        
        for photo in photos:
            # 使用姓名为目录名，空名用ID代替
            dir_name = f"{photo.name or photo.id}"
            
            # 添加原始照片
            if photo.photo and photo.photo.storage.exists(photo.photo.name):
                zipf.writestr(
                    f"{dir_name}/原始照片.jpg",
                    photo.photo.read()
                )
            
            # 添加旋转后照片
            if photo.rotated_photo and photo.rotated_photo.storage.exists(photo.rotated_photo.name):
                zipf.writestr(
                    f"{dir_name}/排版照片.jpg",
                    photo.rotated_photo.read()
                )
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="寸照.zip"'
    return response
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def idcard_add(request):
    title = 'tu'

    if request.method == 'POST':
        # 修改获取参数的name
        front_file = request.FILES.get('人像')  # 原front
        back_file = request.FILES.get('国徽')   # 原back
        # ... 后续代码保持不变 ...
        with Image.open(front_file) as img1:
            img1 = resize_photo(img1, 3)  # 保持原有缩放逻辑

            
            
            img_io = io.BytesIO()            

            img1.save(img_io, format='JPEG')
            img_人像 = img_io.getvalue()  # 直接获取字节数据
            
        with Image.open(back_file) as img2:
            img2 = resize_photo(img2, 3)

            
            
            img_io = io.BytesIO()            
           
            img2.save(img_io, format='JPEG')
            img_国徽 = img_io.getvalue()  # 直接获取字节数据            
            
            
        # 在原有代码中修改

            


        # 合并图片
        combined_img = combine_a4_images(img1, img2)

        # 保存结果
        img_io = io.BytesIO()
        combined_img.save(img_io, format='JPEG', quality=90)
        combined_bytes = img_io.getvalue()
                    
            
            
            
            
            
            
            
            
            
            
            
 
        姓名,身份证号码=sfz(front_file)

        # 创建模型实例
        models.IDCard.objects.create(
            name=姓名,
            id_number=身份证号码,
            
            front_image=ContentFile(img_人像, name=f"{身份证号码}.jpg"),
            back_image=ContentFile(img_国徽, name=f"{身份证号码}_rotated.jpg"),
            combined_image=ContentFile(combined_bytes, name=f"{身份证号码}_双面.jpg")
            
        )


        return redirect("/home/idcard_list")

 
      
 
    
    else:
        model_names = models.IDCard.objects.values_list(
            'name', flat=True).distinct()
        return render(request, 'idcard_add.html', {'model_names': model_names})


def idcard_delete(request):

    id = request.GET.get('id')
    models.IDCard.objects.filter(id=str(id)).delete()
    return redirect("/home/idcard_list")


def idcard_list(request):

    title = 'idcard'
    if request.method == "GET":
        model_fields = models.IDCard._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields]
        cols.append({'verbose_name': '操作'})
        data = models.IDCard.objects.values().order_by('-created_at')[:100]
        return render(request, 'idcard_list.html', {
            "data": data,
            "cols": cols,
            "title": title,
            "export_url": "/home/idcard_export_zip"  # 新增导出URL参数
        })



@资料员
def idcard_export_zip(request):
    import zipfile
    from io import BytesIO
    
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # 获取所有身份证记录
        idcards = models.IDCard.objects.all()
        
        for card in idcards:
            # 创建以身份证号命名的目录
            dir_name = f"{card.id_number}_{card.name}"
            
            # 添加人像面
            if card.front_image and card.front_image.storage.exists(card.front_image.name):
                zipf.writestr(
                    f"{dir_name}/人像面.jpg",
                    card.front_image.read()
                )
            
            # 添加国徽面
            if card.back_image and card.back_image.storage.exists(card.back_image.name):
                zipf.writestr(
                    f"{dir_name}/国徽面.jpg",
                    card.back_image.read()
                )
            
            # 添加合成图片
            if card.combined_image and card.combined_image.storage.exists(card.combined_image.name):
                zipf.writestr(
                    f"{dir_name}/合成双面.jpg",
                    card.combined_image.read()
                )
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="身份证.zip"'
    return response


#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++














def tu_add(request):
    title = 'tu'
    if request.method == 'POST':

        model_name = request.POST['model_name']
        files = request.FILES.getlist('file')  # 获取所有上传的文件

        for file in files:
            # 处理每个文件，例如保存到数据库或文件系统
            models.UploadedTu.objects.create(
                model_name=model_name, pdf_file=file)

        return redirect("/home/tu_list")

    else:
        model_names = models.UploadedTu.objects.values_list(
            'model_name', flat=True).distinct()
        return render(request, 'tu_add.html', {'model_names': model_names})


def tu_delete(request):

    id = request.GET.get('id')
    models.UploadedTu.objects.filter(id=str(id)).delete()
    return redirect("/home/tu_list")


def tu_list(request):

    title = 'tu'
    if request.method == "GET":

        data = models.UploadedTu.objects.values()[:100]

        # print(data)

        lst = dframe(data)
        cols = []

        for i in lst:
            cols.append({'age': i})

        return render(request, 'tu_list.html', {"data": data, "cols": cols, "title": title})

#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
def pdf_add(request):
    title = 'tu'
    if request.method == 'POST':

        model_name = request.POST['model_name']
        files = request.FILES.getlist('file')  # 获取所有上传的文件

        for file in files:
            # 处理每个文件，例如保存到数据库或文件系统
            models.UploadedPDF.objects.create(
                model_name=model_name, pdf_file=file)

        return redirect("/home/pdf_list")

    else:
        model_names = models.UploadedPDF.objects.values_list(
            'model_name', flat=True).distinct()
        return render(request, 'pdf_add.html', {'model_names': model_names})
       

def pdf_delete(request):

    id = request.GET.get('id')
    models.UploadedPDF.objects.filter(id=str(id)).delete()
    return redirect("/home/pdf_list")


def pdf_list(request):

    title = 'pdf'
    if request.method == "GET":

        data = models.UploadedPDF.objects.values()[:100]

        # print(data)

        lst = dframe(data)
        cols = []

        for i in lst:
            cols.append({'age': i})

        return render(request, 'pdf_list.html', {"data": data, "cols": cols, "title": title})


# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################


def inventory_list(request):

    title  = 'inventory'
    database = '出入库记录'
    
    if request.method == "GET":
        
        model_fields = models.ExplosiveInventoryItem._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields]
        cols.append({'verbose_name': '操作'})
        data = models.ExplosiveInventoryItem.objects.values().order_by('-date')[:100]

 

    

        return render(request, 'inventory_list.html', {"data": data, "cols": cols, "数据库": database, 'title': title})


def inventory_add(request):
    title = '出入库记录'

    if request.method == "GET":

        form = ExplosiveInventoryItemForm()

        return render(request, 'create.html', {'form': form, '标题': title})

    if request.method == 'POST':
        form = ExplosiveInventoryItemForm(request.POST)

        if form.is_valid():

            form.save()

        else:

            form.errors
            return render(request, 'create.html', {'form': form, '标题': title})

    return redirect("/home/inventory_list")


def inventory_delete(request):

    id = request.GET.get('id')
    models.ExplosiveInventoryItem.objects.filter(id=str(id)).delete()
    return redirect("/home/inventory_list")


def inventory_edit(request):

    title = '出入库记录'
    id = request.GET.get('id')
    row_object = models.ExplosiveInventoryItem.objects.filter(
        id=str(id)).first()

    if request.method == "GET":

        form = modelform.ExplosiveInventoryItemForm(instance=row_object)

        return render(request, 'create.html', {"form": form, "标题": title})

    form = modelform.ExplosiveInventoryItemForm(
        data=request.POST, instance=row_object)

    if form.is_valid():

        form.save()
    else:
        title = '输入错误'
        form.errors
        return render(request, 'create.html', {'form': form})
    return redirect("/home/inventory_list")


# categorycontent------------------------------------------------------------------------------------------------------------------------------
# categorycontent------------------------------------------------------------------------------------------------------------------------------
# categorycontent------------------------------------------------------------------------------------------------------------------------------
# categorycontent------------------------------------------------------------------------------------------------------------------------------


def categorycontent_list(request):

    database = 'categorycontent'
    title = '民爆物品'
    if request.method == "GET":

        data = models.CategoryContent.objects.values()[:100]

        lst = dframe(data)
        cols = []

        for i in lst:
            cols.append({'age': i})

        return render(request, 'list.html', {"data": data, "cols": cols, "数据库": database, '标题': title})


def categorycontent_create(request):
    title = '民爆物品'

    if request.method == "GET":

        form = modelform.CategoryContentForm()

        return render(request, 'create.html', {'form': form, '标题': title})

    if request.method == 'POST':
        form = modelform. CategoryContentForm(request.POST)

        if form.is_valid():
            form.save()

        else:
            form.errors
            return render(request, 'create.html', {'form': form, '标题': title})

    return redirect("/home/categorycontent_list")


def categorycontent_delete(request):

    id = request.GET.get('id')
    models.CategoryContent.objects.filter(id=str(id)).delete()
    return redirect("/home/categorycontent_list")


def categorycontent_edit(request):
    title = '民爆物品'
    id = request.GET.get('id')
    row_object = models.CategoryContent.objects.filter(id=str(id)).first()

    if request.method == "GET":

        form = modelform.CategoryContentForm(instance=row_object)

        return render(request, 'create.html', {"form": form, "标题": title})

    form = modelform.CategoryContentForm(
        data=request.POST, instance=row_object)

    if form.is_valid():

        form.save()
    else:
        title = '输入错误'
        form.errors
        return render(request, 'create.html', {'form': form})
    return redirect("/home/categorycontent_list")


# ---------------------------------------------------------------------------------------


def upload_model(request):
    # 获取指定应用中的所有模型
    app_models = apps.get_app_config('app01').get_models()
    model_names = [model.__name__ for model in app_models]

    if request.method == 'POST':
        model_name = request.POST.get('model_name')
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            # 处理上传的 '.xlsx' 文件
            df = pd.read_excel(uploaded_file)
            df['password'] = df['password'].apply(lambda x: md5(str(x)))
            # 这里可以根据选择的模型表名进行相应的处理
            print(f"选择的模型表: {model_name}")
            print(df.head())
            try:
                with sl.connect('/Users/sunhongchen/lnjx2025/db.sqlite3') as con:
                    df.to_sql(f'app01_{model_name.lower()}',
                              con, index=False, if_exists='append')

                return HttpResponse("文件上传成功！")
            except Exception as e:
                return HttpResponse(f"文件上传失败: {str(e)}")

    return render(request, 'upload.html', {'model_names': model_names})



#--------------------------------------
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
def candidateprofile_add(request):
    title = 'candidateprofile'
    if request.method == "GET":
        form = modelform.CandidateProfileForm()
        return render(request, 'candidateprofile_add.html', {"form": form, "标题": title})

    form = modelform.CandidateProfileForm(
        data=request.POST, 
        files=request.FILES
    )

    if form.is_valid():
        # 处理上传的证件照
        img_bytes = None
        if 'photo' in request.FILES:
            photo_file = request.FILES['photo']
            with Image.open(photo_file) as img:
                if img.mode in ('RGBA', 'LA'):
                    img = img.convert('RGB')
                
                processed_img = resize_photo(cut_photo(img, 1), 1)
                
                img_io = io.BytesIO()
                processed_img.save(img_io, format='JPEG', quality=90)
                img_bytes = img_io.getvalue()
                
                form.instance.photo.save(
                    f"{form.cleaned_data['name']}_processed.jpg",
                    ContentFile(img_bytes),
                    save=False
                )

        # 保存表单数据到数据库
        instance = form.save()
        
        # 创建Excel文件
 
        wb = Workbook()
        ws = wb.active
        ws.title = "候选人档案"

        # 设置列宽和行高
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35  # 加宽B列适应换行

        # 创建样式对象
        from openpyxl.styles import Font, Alignment
        header_font = Font(size=16)
        cell_alignment = Alignment(vertical='center', wrap_text=True)

        # 写入基础数据
        data_fields = [
            ('姓名', 'name'),
            ('性别', 'gender'),
            ('手机', 'mobile'),
            ('年龄', 'age'),
            ('婚姻状况', 'marital_status'),
            ('学历专业', 'education'),
            ('驾照', 'has_driver_license'),
            ('特长', 'special_skills'),
            ('工作经历', 'work_experience'),
            ('现住所', 'current_address'),
            ('应聘岗位','position'),            
            ('期望薪资', 'expected_salary')
        ]#current_address

        for row, (label, field) in enumerate(data_fields, start=1):
            # 写入A列
            cell = ws[f'A{row}']
            cell.value = label
            cell.font = header_font
            cell.alignment = cell_alignment
            
            # 写入B列
            cell = ws[f'B{row}']
            cell.value = str(form.cleaned_data.get(field, ''))
            cell.font = Font(size=16)
            cell.alignment = cell_alignment
            
            # 设置行高
            ws.row_dimensions[row].height = 35  # 根据字体大小调整行高

        # 插入照片（如果存在）
        if img_bytes:
            img_temp = io.BytesIO(img_bytes)
            img = ExcelImage(img_temp)
            img.width = 110
            img.height = 160
            ws.add_image(img, 'c1')
            
        # 保存Excel文件
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        # 保存到模型的resume_file字段
        instance.resume_file.save(
            f"{form.cleaned_data['name']}_profile.xlsx",
            ContentFile(excel_io.getvalue()),
            save=True
        )

        return HttpResponse("""
            <div style="
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                margin: 0;
            ">
                <h1 style="font-size: 48px; margin: 0">数据上传成功</h1>
            </div>
        """)
        
    else:
        title = '输入错误'
        return render(request, 'candidateprofile_add.html', {"form": form, "标题": title})  
    
    
      
    
    

@资料员
def candidateprofile_delete(request):
    id = request.GET.get('id')
    models.Candidate.objects.filter(id=str(id)).delete()
    return redirect("/home/candidateprofile_list")


def candidateprofile_list(request):
    title = 'candidateprofile'
    if request.method == "GET":
        data = models.Candidate.objects.values().order_by('-created_at')[:100]
        # 获取模型字段的verbose_name
        model_fields = models.Candidate._meta.fields
        cols = [{'verbose_name': 'id'}] +[{'verbose_name': field.verbose_name} for field in model_fields if field.name != 'id']
        
        # 添加操作列
        cols.append({'verbose_name': '操作'})
        
      
        return render(request, 'candidateprofile_list.html', {
            "data": data,
            "cols": cols,
            "title": title
        })



# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
def save_image_to_field(img, filename):
    img_io = io.BytesIO()
    if img.mode in ('RGBA', 'LA'):
        img = img.convert('RGB')
    img.save(img_io, format='JPEG')
    return ContentFile(img_io.getvalue(), name=filename)

def process_photo(img):
    img = cut_photo(img, 1)
    img = resize_photo(img, 1)
    img_io = io.BytesIO()
    img.save(img_io, format='JPEG')
    return ContentFile(img_io.getvalue())

def explosivestaff_add(request):
    title = 'tu'
    if request.method == 'POST':
        if 'zip_file' in request.FILES:
            zip_file = request.FILES['zip_file']
            success_count = 0
            errors = []
            
            try:
                with zipfile.ZipFile(zip_file) as zf:
                    folder_files = defaultdict(dict)
                    for file_info in zf.infolist():
                        # 增强版编码修复（新增多个编码尝试）
                        raw_name = file_info.filename
                        try:
                            # 优先尝试 UTF-8 解码
                            corrected = raw_name.encode('cp437').decode('utf-8')
                        except UnicodeDecodeError:
                            try:
                                # 次尝试 GB18030 解码
                                corrected = raw_name.encode('cp437').decode('gb18030')
                            except:
                                # 最后尝试忽略错误字符
                                corrected = raw_name.encode('cp437').decode('utf-8', 'ignore')
                        
                        # 跳过Mac系统文件（新增过滤）
                        if corrected.startswith('._'):
                            continue
                            
                        if not file_info.is_dir() and '/' in corrected:
                            folder, filename = corrected.split('/', 1)
                            folder_files[folder][filename] = file_info


                    for folder, files in folder_files.items():
                        try:
                            required_files = {
                                '人像.jpg': None,
                                '国徽.jpg': None,
                                '证件照.jpg': None,
                                '无犯罪证明.jpg': None,
                                '毕业证.jpg': None
                            }
                            
                            # 读取文件（添加调试输出）
            
                            for filename in required_files.keys():
                                if filename in files:
                                    try:
                                        file_info = files[filename]
                                        # 添加文件存在性验证
                                        if file_info.file_size == 0:
                                            raise ValueError(f"文件 {filename} 为空")
                                            
                                        with Image.open(io.BytesIO(zf.read(file_info))) as img:
                                            required_files[filename] = img
                                            # 添加图片有效性验证
                                            img.verify()  # 验证图片完整性
                                    except Exception as e:
                                        raise ValueError(f"文件 {filename} 损坏: {str(e)}")
                                else:
                                    raise ValueError(f"缺少文件: {filename}")


     

                        # 处理身份证信息

                        # 添加空值检查
                            if not all(required_files.values()):
                                missing = [k for k, v in required_files.items() if not v]
                                raise ValueError(f"图片加载失败: {missing}")
                                # 处理身份证信息
                                姓名, 身份证号码 = sfz(required_files['人像.jpg'])
                                print(f"姓名: {姓名}, 身份证号码: {身份证号码}")
                                # 处理各图片并保存
                                front_img = resize_photo(required_files['人像.jpg'], 3)
                                back_img = resize_photo(required_files['国徽.jpg'], 3)
                                combined_img = combine_a4_images(front_img, back_img)
                                
                                # 保存到数据库
                                models.ExplosiveStaff.objects.create(
                                    name=姓名,
                                    id_number=身份证号码,
                                    front_image=save_image_to_field(front_img, f"{身份证号码}_front.jpg"),
                                    back_image=save_image_to_field(back_img, f"{身份证号码}_back.jpg"),
                                    combined_image=save_image_to_field(combined_img, f"{身份证号码}_combined.jpg"),
                                    photo=process_photo(required_files['证件照.jpg']),
                                    no_crime=save_image_to_field(required_files['无犯罪证明.jpg'], f"{身份证号码}_no_crime.jpg"),
                                    graduation=save_image_to_field(required_files['毕业证.jpg'], f"{身份证号码}_graduation.jpg")
                                )
                                success_count += 1
                                
                        except Exception as e:
                            errors.append(f"{folder}: {str(e)}")
                            
                return HttpResponse(f"成功导入 {success_count} 条记录，错误 {len(errors)} 条{errors}" )
                
            except zipfile.BadZipFile:
                return HttpResponse("无效的ZIP文件格式", status=400)
                
        else:
            # 原有单个文件处理逻辑保持不变
            front_file = request.FILES.get('front')
            # ... 原有单个文件处理代码 ...
        
            front_file = request.FILES.get('front')
            back_file = request.FILES.get('back')
            photo = request.FILES.get('photo')
            no_crime = request.FILES.get('no_crime')  # 原错误参数 'photo'
            graduation = request.FILES.get('graduation')  # 原错误参数 'photo'  
            mobile = request.POST.get('mobile', '')  # 新增
            bank_card_number= request.POST.get('bank_card_number', '')  # 新增
            
            with Image.open(front_file) as img1:
                img1 = resize_photo(img1, 3)  
                img_io = io.BytesIO()            
                img1.save(img_io, format='JPEG')
                img_人像 = img_io.getvalue()  
                
            with Image.open(back_file) as img2:
                img2 = resize_photo(img2, 3)
                img_io = io.BytesIO()           
                img2.save(img_io, format='JPEG')
                img_国徽 = img_io.getvalue()   
                
                
            with Image.open(photo) as img3:  
                if img3.mode in ('RGBA', 'LA'):
                    img3 = img3.convert('RGB')  # 移除Alpha通道
                img_io = io.BytesIO()
                img3=resize_photo(cut_photo(img3,1),1)
                img3.save(img_io, format='JPEG')
                img_bytes = img_io.getvalue()  # 直接获取字节数据
                rotated_io = io.BytesIO()
                排版(img3,rotated_io)
                rotated_bytes = rotated_io.getvalue()       
            
            # 新增 img4 处理（示例，请根据实际需求修改文件来源）
            with Image.open(no_crime) as img4:  # 替换为实际文件来源如 request.FILES.get('file4')
                img4 = resize_photo(img4, 4)  
                img_io = io.BytesIO()            
                img4.save(img_io, format='JPEG')
                img_数据4 = img_io.getvalue()

            # 新增 img5 处理（示例，请根据实际需求修改文件来源）
            with Image.open(graduation) as img5:  # 替换为实际文件来源如 request.FILES.get('file5')
                img5 = resize_photo(img5, 4)  
                img_io = io.BytesIO()            
                img5.save(img_io, format='JPEG')
                img_数据5 = img_io.getvalue()       
            
            
            
                
                
                
                
                
            # 合并图片逻辑保持不变
            combined_img = combine_a4_images(img1, img2)
            img_io = io.BytesIO()
            combined_img.save(img_io, format='JPEG', quality=90)
            combined_bytes = img_io.getvalue()
    
            try:
                姓名, 身份证号码 = sfz(front_file)
            except:
                姓名 = "未识别"
                身份证号码 = "888888888888888888"  # 18个8

            # 修改模型引用为ExplosiveStaff
            models.ExplosiveStaff.objects.create(
                name=姓名,
                id_number=身份证号码,
                mobile=mobile,  # 新增字段
                bank_card_number
                front_image=ContentFile(img_人像, name=f"{身份证号码}_front.jpg"),
                back_image=ContentFile(img_国徽, name=f"{身份证号码}_back.jpg"),
                combined_image=ContentFile(combined_bytes, name=f"{身份证号码}_combined.jpg"),
                
                photo=ContentFile(img_bytes, name=f"{身份证号码}_photo.jpg"),
                typeset_photo=ContentFile(rotated_bytes, name=f"{身份证号码}_typeset.jpg"), 
                no_crime=ContentFile(img_数据4, name=f"{身份证号码}_no_crime.jpg"), 
                graduation=ContentFile(img_数据5, name=f"{身份证号码}_graduation.jpg"),      
                
            )

            return HttpResponse("""
                <div style="
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100vh;
                    margin: 0;
                ">
                    <h1 style="font-size: 48px; margin: 0">数据上传成功</h1>
                </div>
            """)

    else:
        model_names = models.ExplosiveStaff.objects.values_list(
            'id_number', flat=True).distinct()
        return render(request, 'explosivestaff_add.html', {
            'model_names': model_names,
            'zip_upload': True
        })








def explosivestaff_delete(request):
    id = request.GET.get('id')
    models.ExplosiveStaff.objects.filter(id=str(id)).delete()
    return redirect("/home/explosivestaff_list")  # 修改重定向路径


def explosivestaff_list(request):
    title = 'explosivestaff'
    if request.method == "GET":
        model_fields = models.ExplosiveStaff._meta.fields
        cols = [{'verbose_name': field.verbose_name} for field in model_fields]
        cols.append({'verbose_name': '操作'})
        data = models.ExplosiveStaff.objects.values().order_by('-created_at')[:100]
        return render(request, 'explosivestaff_list.html', {
            "data": data,
            "cols": cols,
            "title": title,
            "export_url": "/home/explosivestaff_export_zip"  # 新增导出URL参数
        })

# 新增的导出函数
def explosivestaff_export_zip(request):

    
    # 创建内存文件
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # 获取所有记录
        staffs = models.ExplosiveStaff.objects.all()
        
        for staff in staffs:
            # 添加所有图片字段到ZIP
            fields = ['front_image', 'back_image', 'combined_image', 
                    'photo', 'typeset_photo', 'no_crime', 'graduation']
            
            for field in fields:
                file = getattr(staff, field)
                if file and file.storage.exists(file.name):
                    zipf.writestr(
                        f"{staff.name}_{staff.id_number}/{field}.jpg",
                        file.read()
                    )
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="explosivestaff_images.zip"'
    return response










# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################
# inventory333333333333333333333333##############################################################################################################

class StaffListView(FilterView, ListView):
    model = models.Admin
    template_name = 'staff_list.html'
    context_object_name = 'data'
    filterset_class = modelform.StaffFilter
    paginate_by = 100  # 每页显示的记录数
    ordering = ['id']  # 默认排序字段

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = '员工列表'
        context['cols'] = [
            {'col_name': 'ID'},
            {'col_name': '标识'},
            {'col_name': '用户名'},
            {'col_name': '身份'},
            {'col_name': '部门'},
            {'col_name': '密码'},
            {'col_name': '图片'},
            {'col_name': '操作'},
            {'col_name': '证件'},
        ]
        return context


def staff_list(request):
    title = '管理员'

    if request.method == "GET":
        data = models.Admin.objects.values()

        # 将 QuerySet 转换为 DataFrame
        df = pd.DataFrame(list(data))

        # 根据 ident 列去重，保留最后一行
        df = df.drop_duplicates(subset='ident', keep='last')
        df=df.sort_values(by='ident', ascending=False)

        # 将 DataFrame 转换回列表
        data = df.to_dict('records')

        lst = dframe(data)

        cols = []
        for i in lst:
            cols.append({'col_name': i})

        return render(request, 'staff_list.html', {"data": data, "cols": cols, "title": title})

    return render(request, 'staff_list.html', {"title": title})

@资料员
def staff_add(request):

    title = '新建员工信息'
    if request.method == "GET":
        form = modelform.Staff()
        return render(request, 'create.html', {"form": form, "标题": title})
    print("request.FILES:", request.FILES) 
    form = modelform.Staff(data=request.POST, files=request.FILES)

    if form.is_valid():
        form.save()
        print("清洗后数据:", form.cleaned_data)  # 注意：is_valid()=False时可能不完整
        print("错误信息:", form.errors.as_json())       
        
    else:
        # 打印原始提交数据
        print("原始提交数据:", form.data)
        
        # 打印每个字段的值
        for field in form:
            print(f"字段 [{field.name}] 值: {field.data} | 错误: {field.errors}")
        
        # 或者更直接的调试方式
        print("清洗后数据:", form.cleaned_data)  # 注意：is_valid()=False时可能不完整
        print("错误信息:", form.errors.as_json())
        
        title = '输入错误'
        
        return render(request, 'create.html', {"form": form, "标题": title})

    return redirect("/staff_list")



@最高权限
def staff_delete(request):

    id = request.GET.get('id')
    models.Admin.objects.filter(id=str(id)).delete()
    return redirect("/staff_list")


@最高权限
def staff_edit(request):
    title = '员工信息编辑'
    id = request.GET.get('id')
    row_object = models.Admin.objects.filter(id=str(id)).first()
    print(models.Admin.objects.filter(id=str(id)).values())
    if request.method == "GET":

        form = modelform.Staff(instance=row_object)

        return render(request, 'change.html', {"form": form, "title": title})

    form = modelform.Staff(
        instance=row_object,  # 先指定要编辑的实例
        data=request.POST,     # 后传入提交数据
        files=request.FILES    # 最后传入文件
    )
    

    if form.is_valid():

        form.save()
    else:
        title = '输入错误'
        form.errors
        return render(request, 'change.html', {"form": form, "title": title})
    return redirect("/staff_list")


#     ⌘ + K → ⌘ + J       # macOS ⌘ + Shift + ]